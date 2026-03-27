import streamlit as st
import json
import pandas as pd
import io
import re
from datetime import datetime
from snowflake.snowpark.context import get_active_session

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

st.set_page_config(layout="wide", page_title="Cortex Project Intel", page_icon="📊")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    :root {
        --bg-page:       #f0f4f1;
        --bg-panel:      #ffffff;
        --bg-panel-alt:  #f7faf8;
        --bg-hover:      #eaf3ee;
        --border:        rgba(34,99,64,0.13);
        --border-strong: rgba(34,99,64,0.26);

        --green-deep:    #0d4428;
        --green-mid:     #196b41;
        --green-accent:  #228b55;
        --green-bright:  #30c472;
        --green-light:   #e4f5ec;
        --green-xlight:  #f0faf4;
        --green-glow:    rgba(48,196,114,0.18);

        --gold:          #c8922a;
        --gold-bg:       #fef7e8;
        --gold-glow:     rgba(200,146,42,0.15);

        --red:           #c0392b;
        --red-bg:        #fdecea;
        --blue:          #1a5ea8;
        --blue-bg:       #e3eeff;
        --purple:        #6b2fa8;
        --purple-bg:     #f0e8ff;
        --teal:          #0e7c6e;
        --teal-bg:       #e4f7f5;

        --text-primary:  #0a1c12;
        --text-secondary:#2a5a3a;
        --text-muted:    #6a9a7a;
        --text-disabled: #a8c8b4;

        --shadow-xs:     0 1px 2px rgba(0,0,0,0.04);
        --shadow-sm:     0 2px 8px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
        --shadow-md:     0 4px 20px rgba(0,0,0,0.08), 0 2px 6px rgba(0,0,0,0.04);
        --shadow-lg:     0 8px 40px rgba(0,0,0,0.10), 0 4px 12px rgba(0,0,0,0.05);
        --shadow-green:  0 4px 24px rgba(34,139,85,0.22);
        --shadow-gold:   0 4px 20px rgba(200,146,42,0.18);

        --radius-xs:     4px;
        --radius-sm:     8px;
        --radius-md:     12px;
        --radius-lg:     18px;
        --radius-xl:     24px;
        --radius-pill:   999px;

        --font-display:  'Syne', sans-serif;
        --font-body:     'Inter', sans-serif;
        --font-mono:     'JetBrains Mono', monospace;
    }

    html, body, [class*="css"], .stApp, div, p, span, label, input, textarea, select, button {
        font-family: var(--font-body) !important;
        font-size: 13px !important;
    }

    .stApp {
        background: var(--bg-page) !important;
        color: var(--text-primary) !important;
    }

    .stApp::before {
        content: '';
        position: fixed;
        inset: 0;
        pointer-events: none;
        z-index: 0;
        background-image:
            radial-gradient(ellipse at 8% 0%, rgba(48,196,114,0.08) 0%, transparent 50%),
            radial-gradient(ellipse at 92% 100%, rgba(34,139,85,0.06) 0%, transparent 45%),
            radial-gradient(ellipse at 50% 50%, rgba(200,146,42,0.03) 0%, transparent 60%);
    }

    .stApp > header { display: none !important; }
    #root > div:first-child { padding-top: 0 !important; }
    .block-container { padding-top: 0.75rem !important; padding-bottom: 2rem !important; max-width: 100% !important; }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #ffffff 0%, #f7faf8 100%) !important;
        border-right: 1.5px solid var(--border) !important;
        box-shadow: 3px 0 20px rgba(0,0,0,0.05) !important;
    }
    [data-testid="stSidebar"] > div:first-child { padding-top: 0 !important; }
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] .stMarkdown p {
        color: var(--text-secondary) !important;
        font-size: 11px !important; font-weight: 500 !important;
    }
    [data-testid="stSidebar"] h3 {
        font-family: var(--font-display) !important;
        font-size: 9px !important; font-weight: 800 !important;
        text-transform: uppercase !important; letter-spacing: 0.20em !important;
        color: var(--green-mid) !important; margin-bottom: 0.6rem !important;
        padding-bottom: 0.4rem !important;
        border-bottom: 1.5px solid var(--border-strong) !important;
    }
    [data-testid="stSidebar"] .stCheckbox label p,
    [data-testid="stSidebar"] .stCheckbox span,
    [data-testid="stSidebar"] .stCheckbox label { color: var(--text-primary) !important; font-size: 12px !important; font-weight: 500 !important; }

    [data-testid="stSidebar"] .stTextInput input,
    [data-testid="stSidebar"] .stTextArea textarea {
        background: #ffffff !important; border: 1.5px solid var(--border-strong) !important;
        border-radius: var(--radius-sm) !important; color: var(--text-primary) !important;
        font-size: 12px !important; box-shadow: var(--shadow-xs) !important;
        transition: all 0.2s !important;
    }
    [data-testid="stSidebar"] .stTextInput input:focus,
    [data-testid="stSidebar"] .stTextArea textarea:focus {
        border-color: var(--green-bright) !important;
        box-shadow: 0 0 0 3px var(--green-glow) !important; outline: none !important;
    }
    [data-testid="stSidebar"] .stSelectbox > div > div,
    [data-testid="stSidebar"] .stMultiSelect > div > div {
        background: #ffffff !important; border: 1.5px solid var(--border-strong) !important;
        border-radius: var(--radius-sm) !important; color: var(--text-primary) !important;
    }
    [data-testid="stSidebar"] .stSlider > div > div > div { background: var(--green-mid) !important; }
    [data-testid="stSidebar"] .stButton > button {
        background: linear-gradient(135deg, var(--green-deep) 0%, var(--green-mid) 50%, var(--green-accent) 100%) !important;
        color: #fff !important; font-family: var(--font-display) !important;
        font-weight: 700 !important; font-size: 11px !important;
        letter-spacing: 0.12em !important; text-transform: uppercase !important;
        border: none !important; border-radius: var(--radius-md) !important;
        padding: 0.75rem 1.2rem !important; box-shadow: var(--shadow-green) !important;
        transition: all 0.25s cubic-bezier(0.34,1.56,0.64,1) !important;
        width: 100% !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: linear-gradient(135deg, var(--green-mid) 0%, var(--green-bright) 100%) !important;
        box-shadow: 0 8px 32px rgba(48,196,114,0.38) !important;
        transform: translateY(-2px) scale(1.01) !important;
    }
    [data-testid="stSidebar"] .stNumberInput input {
        background: #ffffff !important; border: 1.5px solid var(--border-strong) !important;
        color: var(--text-primary) !important; border-radius: var(--radius-sm) !important;
    }

    h1 { font-family: var(--font-display) !important; font-size: 22px !important; font-weight: 800 !important; color: var(--text-primary) !important; margin: 0 !important; }
    h2 { font-family: var(--font-display) !important; font-size: 15px !important; font-weight: 700 !important; color: var(--text-primary) !important; margin: 0 0 0.9rem 0 !important; }
    h3 { font-family: var(--font-display) !important; font-size: 13px !important; font-weight: 600 !important; color: var(--text-primary) !important; margin-bottom: 0.6rem !important; }

    [data-testid="stMetric"] {
        background: #ffffff !important;
        border-radius: var(--radius-md) !important;
        padding: 1rem 1.1rem !important;
        border: 1.5px solid var(--border) !important;
        box-shadow: var(--shadow-sm) !important;
        transition: all 0.22s !important;
        position: relative !important;
        overflow: hidden !important;
    }
    [data-testid="stMetric"]::before {
        content: '';
        position: absolute; top: 0; left: 0; right: 0; height: 2.5px;
        background: linear-gradient(90deg, var(--green-mid), var(--green-bright), var(--gold));
    }
    [data-testid="stMetric"]:hover {
        border-color: var(--border-strong) !important;
        box-shadow: var(--shadow-green) !important;
        transform: translateY(-1px) !important;
    }
    [data-testid="stMetricValue"] { font-family: var(--font-display) !important; font-size: 20px !important; font-weight: 800 !important; color: var(--text-primary) !important; }
    [data-testid="stMetricLabel"] { font-size: 9px !important; font-weight: 700 !important; color: var(--text-muted) !important; text-transform: uppercase !important; letter-spacing: 0.10em !important; }
    [data-testid="stMetricDelta"] { font-size: 10px !important; }

    .stButton > button {
        background: var(--green-accent) !important;
        color: #fff !important; font-family: var(--font-display) !important;
        font-weight: 700 !important; font-size: 11px !important;
        letter-spacing: 0.09em !important; text-transform: uppercase !important;
        border: none !important; border-radius: var(--radius-sm) !important;
        padding: 0.6rem 1.4rem !important;
        box-shadow: var(--shadow-green) !important;
        transition: all 0.2s cubic-bezier(0.34,1.56,0.64,1) !important;
        width: 100% !important;
    }
    .stButton > button:hover {
        background: var(--green-bright) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(48,196,114,0.32) !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        background: #ffffff !important;
        border-radius: var(--radius-md) !important;
        border: 1.5px solid var(--border) !important;
        padding: 0.3rem !important;
        gap: 0.2rem !important;
        box-shadow: var(--shadow-sm) !important;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        color: var(--text-secondary) !important;
        font-family: var(--font-display) !important;
        font-size: 11px !important; font-weight: 700 !important;
        border-radius: var(--radius-sm) !important;
        padding: 0.5rem 0.95rem !important;
        letter-spacing: 0.04em !important;
        transition: all 0.18s !important;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, var(--green-mid), var(--green-accent)) !important;
        color: #fff !important;
        box-shadow: 0 2px 10px rgba(34,139,85,0.3) !important;
    }
    .stTabs [data-baseweb="tab-panel"] {
        background: #ffffff !important;
        border: 1.5px solid var(--border) !important;
        border-radius: 0 0 var(--radius-lg) var(--radius-lg) !important;
        padding: 1.4rem !important;
        margin-top: -1px !important;
        box-shadow: var(--shadow-sm) !important;
    }

    .streamlit-expanderHeader {
        background: #ffffff !important;
        border-radius: var(--radius-sm) !important;
        border: 1.5px solid var(--border) !important;
        padding: 0.65rem 1rem !important;
        font-family: var(--font-display) !important;
        font-weight: 700 !important; font-size: 12px !important;
        color: var(--text-secondary) !important;
    }
    .streamlit-expanderContent {
        background: var(--bg-panel-alt) !important;
        border: 1.5px solid var(--border) !important;
        border-top: none !important;
        border-radius: 0 0 var(--radius-sm) var(--radius-sm) !important;
        padding: 1rem !important;
    }

    .stAlert {
        background: var(--green-xlight) !important;
        border: 1.5px solid var(--border-strong) !important;
        border-left: 3px solid var(--green-bright) !important;
        border-radius: var(--radius-md) !important;
        color: var(--text-primary) !important; font-size: 12px !important;
    }

    .stTextArea textarea, .stTextInput input {
        background: #ffffff !important; color: var(--text-primary) !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius-sm) !important; font-size: 12px !important;
        transition: all 0.18s !important;
    }
    .stTextArea textarea:focus, .stTextInput input:focus {
        border-color: var(--green-bright) !important;
        box-shadow: 0 0 0 3px var(--green-glow) !important;
    }
    .stSelectbox > div > div, .stMultiSelect > div > div {
        background: #ffffff !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius-sm) !important;
        color: var(--text-primary) !important; font-size: 12px !important;
    }
    .stSlider > div > div > div { background: var(--green-accent) !important; }
    .stDataFrame { border: 1.5px solid var(--border) !important; border-radius: var(--radius-md) !important; font-size: 12px !important; }
    hr { border: none !important; height: 1px !important; background: var(--border) !important; margin: 1.5rem 0 !important; }

    .top-header {
        position: sticky; top: 0; z-index: 999;
        background: rgba(255,255,255,0.97);
        backdrop-filter: blur(24px) saturate(180%);
        -webkit-backdrop-filter: blur(24px) saturate(180%);
        border-bottom: 1.5px solid var(--border);
        padding: 0.65rem 1.5rem;
        margin: -0.75rem -1rem 1.5rem -1rem;
        display: flex; align-items: center; gap: 1rem;
        box-shadow: 0 2px 20px rgba(0,0,0,0.06);
    }

    .live-pill {
        display: inline-flex; align-items: center; gap: 0.4rem;
        background: var(--green-light);
        border: 1.5px solid var(--border-strong);
        border-radius: var(--radius-pill);
        padding: 0.28rem 0.75rem;
        font-size: 10px; font-weight: 700;
        font-family: var(--font-display) !important;
        color: var(--green-deep);
        letter-spacing: 0.08em; text-transform: uppercase;
    }
    .live-dot { width: 7px; height: 7px; background: var(--green-bright); border-radius: 50%; position: relative; flex-shrink: 0; }
    .live-dot::after {
        content: ''; position: absolute; inset: -3px; border-radius: 50%;
        border: 2px solid rgba(48,196,114,0.5);
        animation: live-ping 1.5s ease-out infinite;
    }
    @keyframes live-ping { 0%{transform:scale(0.8);opacity:1;} 100%{transform:scale(2.4);opacity:0;} }

    .tag {
        display: inline-block;
        background: var(--green-xlight);
        border: 1.5px solid var(--border);
        color: var(--green-deep);
        font-size: 9px; font-weight: 700;
        font-family: var(--font-display) !important;
        padding: 0.15rem 0.6rem;
        border-radius: var(--radius-xs);
        margin-right: 0.3rem;
        letter-spacing: 0.06em;
    }

    .panel {
        background: #ffffff;
        border: 1.5px solid var(--border);
        border-radius: var(--radius-lg);
        padding: 1.35rem 1.5rem;
        margin-bottom: 1.1rem;
        box-shadow: var(--shadow-sm);
        position: relative; overflow: hidden;
        transition: box-shadow 0.22s;
    }
    .panel::before {
        content: '';
        position: absolute; top: 0; left: 0; right: 0; height: 2.5px;
        background: linear-gradient(90deg, var(--green-mid), var(--green-bright), var(--gold), transparent);
    }
    .panel:hover { box-shadow: var(--shadow-md); }

    .section-label {
        font-family: var(--font-display) !important;
        font-size: 9px; font-weight: 800;
        text-transform: uppercase; letter-spacing: 0.18em;
        color: var(--green-mid); margin-bottom: 0.35rem; opacity: 0.75;
    }

    .sb-card {
        background: #ffffff;
        border: 1.5px solid var(--border);
        border-radius: var(--radius-md);
        padding: 0.9rem 1rem; margin-bottom: 0.75rem;
        box-shadow: var(--shadow-xs);
        transition: box-shadow 0.2s, border-color 0.2s;
    }
    .sb-card:hover { border-color: var(--border-strong) !important; box-shadow: var(--shadow-sm) !important; }

    .proj-card {
        background: #ffffff;
        border: 1.5px solid var(--border);
        border-radius: var(--radius-md);
        padding: 0.9rem 1rem; margin-bottom: 0.5rem;
        transition: border-color 0.2s, box-shadow 0.2s;
        box-shadow: var(--shadow-xs);
    }
    .proj-card:hover { border-color: var(--border-strong); box-shadow: var(--shadow-green); transform: translateY(-1px); }

    .sprint-card {
        background: #ffffff;
        border: 1.5px solid var(--border-strong);
        border-radius: var(--radius-md);
        padding: 0.9rem 1rem; margin-bottom: 0.6rem;
        border-left: 3px solid var(--green-mid);
        box-shadow: var(--shadow-xs);
    }

    .pulse-loader {
        display: flex; flex-direction: column; align-items: center; justify-content: center;
        gap: 1.4rem; padding: 3rem 2rem;
        background: #ffffff;
        border: 1.5px solid var(--border);
        border-radius: var(--radius-lg); margin-bottom: 1.1rem;
        box-shadow: var(--shadow-md);
    }
    .pulse-rings { position: relative; width: 54px; height: 54px; display: flex; align-items: center; justify-content: center; }
    .pulse-rings span { position: absolute; border-radius: 50%; border: 2px solid var(--green-mid); animation: pulse-expand 2s ease-out infinite; }
    .pulse-rings span:nth-child(1){width:54px;height:54px;animation-delay:0s;}
    .pulse-rings span:nth-child(2){width:38px;height:38px;animation-delay:0.35s;}
    .pulse-rings span:nth-child(3){width:22px;height:22px;animation-delay:0.7s;}
    .pulse-rings span:nth-child(4){width:10px;height:10px;background:var(--green-mid);border:none;animation:none;}
    @keyframes pulse-expand { 0%{transform:scale(0.4);opacity:1;} 100%{transform:scale(1);opacity:0;} }
    .pulse-label { font-family: var(--font-display) !important; font-size: 12px; font-weight: 800; color: var(--text-secondary); letter-spacing: 0.10em; text-transform: uppercase; text-align: center; }
    .pulse-sublabel { font-size: 11px; color: var(--text-muted); text-align: center; margin-top: -0.8rem; }
    .pulse-steps { display: flex; flex-direction: column; gap: 0.5rem; width: 100%; max-width: 340px; }
    .pulse-step { display: flex; align-items: center; gap: 0.6rem; font-size: 11px; color: var(--text-disabled); }
    .pulse-step.active { color: var(--text-primary); }
    .step-dot { width: 6px; height: 6px; border-radius: 50%; background: var(--border); flex-shrink: 0; }
    .pulse-step.active .step-dot { background: var(--green-mid); box-shadow: 0 0 6px rgba(34,139,85,0.5); animation: blink 0.9s ease-in-out infinite; }
    @keyframes blink { 0%,100%{opacity:1;} 50%{opacity:0.3;} }

    .plan-section-header {
        font-family: var(--font-display) !important;
        font-size: 11px !important; font-weight: 800 !important;
        color: var(--green-mid) !important;
        text-transform: uppercase !important; letter-spacing: 0.14em !important;
        margin: 1.4rem 0 0.5rem 0 !important; padding-bottom: 0.4rem !important;
        border-bottom: 2px solid var(--border-strong) !important; display: block !important;
    }
    .plan-table-wrap {
        overflow-x: auto !important; margin: 0.6rem 0 1.5rem 0 !important;
        border-radius: var(--radius-sm) !important;
        border: 1.5px solid var(--border-strong) !important;
        box-shadow: var(--shadow-sm) !important;
    }
    .plan-table { width: 100% !important; border-collapse: collapse !important; font-size: 11.5px !important; }
    .plan-th {
        background: var(--green-xlight) !important; color: var(--green-deep) !important;
        padding: 0.6rem 1rem !important; text-align: left !important;
        font-family: var(--font-display) !important;
        font-size: 10px !important; font-weight: 800 !important;
        letter-spacing: 0.10em !important; text-transform: uppercase !important;
        border-bottom: 2px solid var(--border-strong) !important;
        border-right: 1px solid var(--border) !important;
    }
    .plan-td {
        padding: 0.5rem 1rem !important; font-size: 11px !important;
        border-bottom: 1px solid var(--border) !important;
        border-right: 1px solid var(--border) !important;
        vertical-align: top !important; line-height: 1.65 !important;
        color: var(--text-primary) !important;
    }

    .qa-question {
        background: var(--green-xlight) !important;
        border: 1.5px solid var(--border-strong) !important;
        border-left: 3px solid var(--green-mid) !important;
        border-radius: var(--radius-sm) !important;
        padding: 0.75rem 1rem !important; margin-bottom: 0.35rem !important;
        font-size: 12px !important; color: var(--green-deep) !important;
        font-weight: 600 !important; display: block !important;
    }
    .qa-answer {
        background: var(--bg-panel-alt) !important;
        border: 1.5px solid var(--border) !important;
        border-radius: 0 var(--radius-sm) var(--radius-sm) var(--radius-sm) !important;
        padding: 0.9rem 1rem !important; margin-bottom: 1rem !important;
        font-size: 12px !important; color: var(--text-primary) !important;
        line-height: 1.8 !important; display: block !important;
    }

    .plan-preview-wrap {
        background: var(--bg-panel-alt) !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius-sm) !important;
        padding: 1rem 1.25rem !important;
        max-height: 320px !important; overflow-y: auto !important;
    }
    .plan-preview-wrap pre {
        font-family: var(--font-mono) !important;
        font-size: 11px !important; color: var(--text-primary) !important;
        line-height: 1.75 !important; white-space: pre-wrap !important;
        word-break: break-word !important; margin: 0 !important; background: transparent !important;
    }

    .csv-badge-imported {
        display: inline-block !important; background: var(--green-light) !important;
        border: 1.5px solid var(--border-strong) !important; color: var(--green-deep) !important;
        font-size: 9px !important; font-weight: 700 !important; padding: 0.1rem 0.45rem !important;
        border-radius: var(--radius-xs) !important; text-transform: uppercase !important; margin-left: 0.4rem !important;
    }
    .csv-badge-pending {
        display: inline-block !important; background: var(--gold-bg) !important;
        border: 1.5px solid rgba(200,146,42,0.35) !important; color: var(--gold) !important;
        font-size: 9px !important; font-weight: 700 !important; padding: 0.1rem 0.45rem !important;
        border-radius: var(--radius-xs) !important; text-transform: uppercase !important; margin-left: 0.4rem !important;
    }

    textarea:disabled, textarea[disabled] {
        color: var(--text-primary) !important;
        -webkit-text-fill-color: var(--text-primary) !important;
        opacity: 1 !important; background: var(--bg-panel-alt) !important;
    }

    .roi-card {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        border: 1.5px solid rgba(34,197,94,0.28);
        border-radius: var(--radius-lg);
        padding: 1.1rem 1.25rem; margin-bottom: 1rem;
        box-shadow: var(--shadow-green);
        position: relative; overflow: hidden;
    }
    .roi-card::before {
        content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2.5px;
        background: linear-gradient(90deg, #16a34a, #22c55e, #4ade80);
    }
    .roi-number { font-family: var(--font-display) !important; font-size: 28px !important; font-weight: 900 !important; color: #15803d !important; line-height: 1 !important; letter-spacing: -0.02em !important; }
    .roi-label { font-family: var(--font-display) !important; font-size: 9px !important; font-weight: 800 !important; color: #166534 !important; text-transform: uppercase !important; letter-spacing: 0.16em !important; }

    .plan-section-title {
        font-family: var(--font-display) !important;
        font-size: 11px; font-weight: 700; color: var(--green-mid);
        text-transform: uppercase; letter-spacing: 0.12em; margin-bottom: 0.5rem;
    }

    .footer {
        text-align: center; margin-top: 3rem; padding: 1.5rem;
        color: var(--text-muted); font-size: 11px; letter-spacing: 0.05em;
        border-top: 1.5px solid var(--border);
    }
    .footer span { color: var(--green-mid); font-weight: 700; font-family: var(--font-display) !important; }

    .health-widget {
        background: #ffffff; border: 1.5px solid var(--border);
        border-radius: var(--radius-md); padding: 0.9rem 1rem; margin-bottom: 0.75rem;
        box-shadow: var(--shadow-xs);
    }
    .effort-card {
        background: var(--green-xlight); border: 1.5px solid var(--border-strong);
        border-radius: var(--radius-md); padding: 0.85rem 1rem; margin-bottom: 0.75rem;
        box-shadow: var(--shadow-xs);
    }
    .ai-tip-card {
        background: var(--gold-bg); border: 1.5px solid rgba(200,146,42,0.22);
        border-left: 4px solid var(--gold);
        border-radius: 0 var(--radius-md) var(--radius-md) 0;
        padding: 0.8rem 0.9rem; margin-bottom: 0.75rem;
    }

    .sprint-board-header {
        background: linear-gradient(135deg, #0d3d22 0%, #196b41 100%);
        border-radius: var(--radius-md) var(--radius-md) 0 0;
        padding: 0.9rem 1.1rem;
        color: #fff;
        font-family: var(--font-display) !important;
        font-weight: 700; font-size: 12px;
        display: flex; align-items: center; gap: 0.6rem;
    }
    .sprint-task-row {
        display: flex; gap: 0.5rem; padding: 0.3rem 0;
        border-bottom: 1px solid var(--border);
        font-size: 11px; color: var(--text-secondary);
        align-items: flex-start;
    }
    .sprint-task-row:last-child { border-bottom: none; }

    .plan-content-wrap {
        background: #ffffff;
        border: 1.5px solid var(--border);
        border-radius: var(--radius-lg);
        padding: 1.75rem 2rem;
        box-shadow: var(--shadow-sm);
    }

    .risk-heatmap-badge {
        display: inline-flex; align-items: center; justify-content: center;
        width: 60px; height: 28px;
        border-radius: var(--radius-xs);
        font-size: 10px; font-weight: 700;
        font-family: var(--font-display) !important;
        letter-spacing: 0.05em;
    }
    .risk-critical { background: #fdecea; color: #c0392b; border: 1.5px solid #c0392b44; }
    .risk-high     { background: #fff0e8; color: #c0682b; border: 1.5px solid #c0682b44; }
    .risk-medium   { background: var(--gold-bg); color: var(--gold); border: 1.5px solid rgba(200,146,42,0.35); }
    .risk-low      { background: var(--green-xlight); color: var(--green-mid); border: 1.5px solid var(--border-strong); }

    .sp-badge {
        display: inline-flex; align-items: center; justify-content: center;
        width: 28px; height: 20px;
        background: var(--blue-bg); color: var(--blue);
        border: 1.5px solid rgba(26,94,168,0.25);
        border-radius: var(--radius-xs);
        font-size: 10px; font-weight: 800;
        font-family: var(--font-display) !important;
    }

    .insight-card {
        background: linear-gradient(135deg, #f8fbf9, #f0faf4);
        border: 1.5px solid var(--border-strong);
        border-radius: var(--radius-md);
        padding: 0.85rem 1rem; margin-bottom: 0.65rem;
        position: relative; overflow: hidden;
    }
    .insight-card::before {
        content: '';
        position: absolute; left: 0; top: 0; bottom: 0; width: 3px;
        background: var(--green-bright);
    }
    .insight-number {
        font-family: var(--font-display) !important;
        font-size: 24px; font-weight: 900; color: var(--green-mid);
        line-height: 1; letter-spacing: -0.02em;
    }

    .velocity-bar {
        height: 8px; border-radius: var(--radius-pill);
        background: var(--border); overflow: hidden; margin-bottom: 0.5rem;
    }
    .velocity-fill {
        height: 100%; border-radius: var(--radius-pill);
        background: linear-gradient(90deg, var(--green-mid), var(--green-bright));
        transition: width 0.6s cubic-bezier(0.34,1.56,0.64,1);
    }

    .dep-badge {
        display: inline-block;
        background: var(--purple-bg); color: var(--purple);
        border: 1.5px solid rgba(107,47,168,0.2);
        font-size: 9px; font-weight: 700;
        padding: 0.1rem 0.5rem; border-radius: var(--radius-xs);
        font-family: var(--font-display) !important;
        letter-spacing: 0.06em; text-transform: uppercase;
    }

    .stakeholder-cell {
        padding: 0.5rem 0.75rem;
        border-radius: var(--radius-sm);
        font-size: 11px; font-weight: 600;
        text-align: center;
    }

    .progress-ring-wrap {
        display: flex; align-items: center; gap: 1rem;
        padding: 0.75rem 0;
    }

    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(8px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    .panel, [data-testid="stMetric"] {
        animation: fadeInUp 0.35s ease both;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="top-header">
    <div style="width:34px;height:34px;
                background:linear-gradient(135deg,#196b41,#0d4428);
                border-radius:10px;display:flex;align-items:center;justify-content:center;
                font-size:1rem;box-shadow:0 3px 14px rgba(48,196,114,0.28);flex-shrink:0;">📊</div>
    <div style="display:flex;flex-direction:column;">
        <div style="font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:0.20em;color:#196b41;line-height:1;font-family:'Syne',sans-serif;">Snowflake Cortex</div>
        <div style="font-size:16px;font-weight:800;color:#0a1c12;letter-spacing:-0.02em;line-height:1.2;font-family:'Syne',sans-serif;">Cortex Project Intel</div>
    </div>
    <div style="font-size:11px;color:#6a9a7a;margin-top:5px;">AI-Powered Project Planning from Institutional Memory</div>
    <div style="margin-left:auto;display:flex;gap:0.5rem;align-items:center;">
        <span class="live-pill"><span class="live-dot"></span>Live</span>
        <span class="tag">Cortex AI</span>
        <span class="tag">PMO Intelligence</span>
        <span class="tag">COCO 2026</span>
    </div>
</div>
""", unsafe_allow_html=True)

session = get_active_session()

def cortex_call(prompt_text, label="Cortex"):
    safe = prompt_text.replace("'", "''")
    if len(safe) > 14000:
        safe = safe[:14000]
    try:
        result = session.sql("SELECT cortex_complete('" + safe + "') as r").collect()[0][0]
        return result or ""
    except Exception as e:
        err = str(e)
        if "400" in err or "token" in err.lower() or "limit" in err.lower():
            st.warning("Trimming " + label + " prompt and retrying…")
            try:
                result = session.sql(
                    "SELECT cortex_complete('" + safe[:7000] + "') as r"
                ).collect()[0][0]
                return result or ""
            except Exception as e2:
                st.error(label + " failed after retry: " + str(e2))
                return ""
        st.error(label + " error: " + str(e))
        return ""

# Session state initialisation
for k, v in [("plan", None), ("plan_name", ""), ("qa_history", []),
             ("csv_projects", []), ("past_projects_df", None),
             ("wbs_structured", None), ("plan_methodology", "Agile / Scrum"),
             ("sow_requirements", None), ("past_patterns", None),
             ("csv_imported", False), ("imported_csv_names", set()),
             ("risk_lessons", None), ("stakeholder_map", None),
             ("release_notes", None), ("velocity_data", None),
             ("decision_log", None), ("current_risks", [])]:
    if k not in st.session_state:
        st.session_state[k] = v

def safe_str(val):
    return '' if pd.isna(val) else str(val)

def parse_date_safe(date_str):
    if not date_str: return ''
    try:
        dt = pd.to_datetime(date_str, errors='coerce')
        if pd.notna(dt): return dt.strftime('%Y-%m-%d')
    except Exception: pass
    return ''

def load_past_projects():
    try:
        df = session.sql("SELECT * FROM projects").to_pandas()
        if df.empty:
            return pd.DataFrame(columns=['project_id','project_name','description',
                                         'wbs_summary','risk_log_summary','test_cases_summary','deployment_plan'])
        id_candidates = ['project_id','id','projectid','PROJECT_ID','PROJECTID']
        found_id = next((c for c in id_candidates if c in df.columns), None)
        if found_id: df.rename(columns={found_id: 'project_id'}, inplace=True)
        else: df['project_id'] = df.index + 1
        for col in ['project_name','description','wbs_summary','risk_log_summary','test_cases_summary','deployment_plan']:
            if col not in df.columns: df[col] = ''
        df = df.drop_duplicates(subset=['project_name'])
        return df[['project_id','project_name','description','wbs_summary',
                   'risk_log_summary','test_cases_summary','deployment_plan']]
    except Exception as e:
        st.error(f"Error loading projects: {e}")
        return pd.DataFrame(columns=['project_id','project_name','description',
                                     'wbs_summary','risk_log_summary','test_cases_summary','deployment_plan'])

def parse_csv_as_project(uploaded_file):
    try:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.strip()
        col_map = {}
        for col in df.columns:
            cl = col.lower()
            if 'task' in cl and 'sub' not in cl: col_map['tasks'] = col
            elif 'subtask' in cl or 'sub task' in cl: col_map['subtasks'] = col
            elif 'start date' in cl: col_map['start_date'] = col
            elif 'end date' in cl: col_map['end_date'] = col
            elif 'project name' in cl: col_map['project_name'] = col
        project_name = safe_str(df[col_map['project_name']].iloc[0]) if 'project_name' in col_map \
                       else uploaded_file.name.replace('.csv','').replace('_',' ').title()
        wbs_lines = []
        for _, row in df.iterrows():
            task    = safe_str(row.get(col_map.get('tasks',''), ''))
            subtask = safe_str(row.get(col_map.get('subtasks',''), ''))
            start   = safe_str(row.get(col_map.get('start_date',''), ''))
            end     = safe_str(row.get(col_map.get('end_date',''), ''))
            if task or subtask:
                line = f"- {task}"
                if subtask: line += f" → {subtask}"
                if start or end: line += f" (Start: {start}, End: {end})"
                wbs_lines.append(line)
        start_date, end_date = '', ''
        if 'start_date' in col_map:
            for val in df[col_map['start_date']]:
                if pd.notna(val) and val: start_date = parse_date_safe(safe_str(val)); break
        if 'end_date' in col_map:
            for val in df[col_map['end_date']]:
                if pd.notna(val) and val: end_date = parse_date_safe(safe_str(val)); break
        return {"project_name": project_name,
                "description": f"Imported from {uploaded_file.name}. {len(df)} rows.",
                "wbs_summary": "\n".join(wbs_lines) or "No tasks found.",
                "start_date": start_date, "end_date": end_date, "technologies_used": ''}
    except Exception as e:
        st.error(f"Error parsing CSV: {e}")
        return None

def parse_wbs_into_sprints(plan_text, duration_weeks, methodology):
    is_agile = any(k in methodology.lower() for k in ['agile','scrum','safe'])
    sprint_len = 2
    num_sprints = max(1, duration_weeks // sprint_len)
    unit = "Sprint" if is_agile else "Phase"

    prompt = (
        f"You are a PMO analyst. Read the project plan below and extract every sprint or phase "
        f"into a JSON array. Use the ACTUAL task names and deliverables from the plan — "
        f"do NOT invent generic placeholder text.\n\n"
        f"Plan text:\n{plan_text[:4000]}\n\n"
        f"Rules:\n"
        f"- Extract {num_sprints} {unit}s covering the full {duration_weeks}-week timeline.\n"
        f"- 'stories' must contain the REAL tasks from the plan (4-6 tasks per sprint).\n"
        f"- 'goal' must be the REAL sprint objective from the plan.\n"
        f"- 'deliverable' must be the REAL acceptance criterion or output.\n"
        f"- Copy task names verbatim from the plan — no paraphrasing, no placeholders.\n\n"
        f"Return ONLY a valid JSON array, no markdown fences, no extra text:\n"
        f'[{{"sprint_num":1,"name":"{unit} 1: <real title>","weeks":"Week 1-2",'
        f'"goal":"<real goal from plan>","stories":["<real task 1>","<real task 2>","<real task 3>","<real task 4>"],'
        f'"deliverable":"<real deliverable from plan>"}}]'
    )
    try:
        safe = prompt.replace("'", "''")
        raw  = session.sql(f"SELECT cortex_complete('{safe}') as r").collect()[0][0]
        raw = raw.strip()
        for fence in ["```json", "```JSON", "```"]:
            if raw.startswith(fence): raw = raw[len(fence):]
            if raw.endswith(fence):   raw = raw[:-len(fence)]
        raw = raw.strip()
        start = raw.find('['); end = raw.rfind(']')
        if start != -1 and end != -1: raw = raw[start:end+1]
        parsed = json.loads(raw)
        if isinstance(parsed, list) and len(parsed) > 0:
            first_stories = parsed[0].get('stories', [])
            placeholders = {'define','build','test','task 1','task 2','task 3','<real task 1>'}
            real_tasks = [s for s in first_stories if str(s).lower().strip() not in placeholders and len(str(s)) > 10]
            if real_tasks:
                return parsed
    except Exception:
        pass

    sprints_from_timeline = {}
    tl_section = re.search(
        r'##\s*DELIVERY TIMELINE(.*?)(?=##\s+[A-Z]|\Z)',
        plan_text, re.DOTALL | re.IGNORECASE
    )
    if tl_section:
        for line in tl_section.group(1).split('\n'):
            parts = [p.strip() for p in line.split('|') if p.strip()]
            if len(parts) >= 3 and not all(set(p) <= set('-: ') for p in parts):
                skip = {'week','sprint','phase','milestone','owner','status','risk','flag','---'}
                if parts[0].lower() in skip: continue
                week_lbl    = parts[0]
                sprint_name = parts[1] if len(parts) > 1 else f"{unit} {len(sprints_from_timeline)+1}"
                milestone   = parts[2] if len(parts) > 2 else ""
                key = sprint_name
                if key not in sprints_from_timeline:
                    sprints_from_timeline[key] = {
                        "sprint_num": len(sprints_from_timeline) + 1,
                        "name": sprint_name, "weeks": week_lbl, "goal": milestone,
                        "stories": [], "deliverable": milestone
                    }
                else:
                    if milestone and milestone not in sprints_from_timeline[key]['stories']:
                        sprints_from_timeline[key]['stories'].append(milestone)

    if sprints_from_timeline:
        result = list(sprints_from_timeline.values())
        for sp in result:
            if not sp['stories'] and sp['goal']:
                sp['stories'] = [sp['goal']]
        return result

    sprints_parsed = []
    current_sprint = None
    sprint_pattern = re.compile(
        r'^(?:#{1,3}\s*)?(' + unit + r'\s*\d+[:\-–]?\s*.+)$', re.IGNORECASE
    )
    week_pattern = re.compile(r'[Ww]eek\s*\d+', re.IGNORECASE)

    for line in plan_text.split('\n'):
        s = line.strip()
        if not s: continue
        m = sprint_pattern.match(s)
        if m:
            if current_sprint and current_sprint['stories']:
                sprints_parsed.append(current_sprint)
            idx = len(sprints_parsed) + 1
            wk_m = week_pattern.search(s)
            current_sprint = {
                "sprint_num": idx, "name": s.lstrip('#').strip(),
                "weeks": wk_m.group(0) if wk_m else f"Week {(idx-1)*sprint_len+1}–{idx*sprint_len}",
                "goal": "", "stories": [], "deliverable": ""
            }
        elif current_sprint:
            if re.match(r'^[-*•🔵▸]\s+', s):
                task = re.sub(r'^[-*•🔵▸]\s+', '', s).strip()
                if task and len(task) > 5: current_sprint['stories'].append(task)
            elif re.match(r'^\d+\.\s+', s):
                task = re.sub(r'^\d+\.\s+', '', s).strip()
                if task and len(task) > 5: current_sprint['stories'].append(task)
            elif not current_sprint['goal'] and len(s) > 10 and not s.startswith('#'):
                current_sprint['goal'] = s
    if current_sprint and current_sprint['stories']:
        sprints_parsed.append(current_sprint)

    if sprints_parsed:
        return sprints_parsed

    deliverables = []
    for line in plan_text.split('\n'):
        s = line.strip()
        m = re.match(r'^\d+\.\s+(.+)', s)
        if m and len(m.group(1)) > 15:
            deliverables.append(m.group(1))
        elif re.match(r'^[-*•]\s+', s):
            task = re.sub(r'^[-*•]\s+', '', s).strip()
            if len(task) > 20: deliverables.append(task)
    deliverables = list(dict.fromkeys(deliverables))[:num_sprints * 5]

    sprints = []
    tasks_per_sprint = max(3, len(deliverables) // max(num_sprints, 1))
    for i in range(num_sprints):
        w_start = i * sprint_len + 1; w_end = (i + 1) * sprint_len
        chunk   = deliverables[i*tasks_per_sprint:(i+1)*tasks_per_sprint]
        if not chunk: chunk = [f"Deliver {unit.lower()} {i+1} objectives"]
        label = f"{unit} {i+1}"
        phase_names = ["Discovery & Setup", "Architecture & Design", "Build & Integrate",
                       "Test & Validate", "UAT & Sign-off", "Deploy & Handover"]
        phase_label = phase_names[min(i, len(phase_names)-1)]
        sprints.append({
            "sprint_num": i+1, "name": f"{label}: {phase_label}",
            "weeks": f"Week {w_start}–{w_end}", "goal": chunk[0] if chunk else f"Complete {phase_label}",
            "stories": chunk, "deliverable": chunk[-1] if chunk else f"{phase_label} complete"
        })
    return sprints


def build_excel(plan_text, plan_name, wbs_sprints, methodology):
    import openpyxl, re as _re
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Helper to convert color to aRGB (8 hex digits)
    def to_aRGB(color):
        if not color:
            return "FF000000"
        color = str(color).strip()
        if color.startswith('#'):
            color = color[1:]
        if len(color) == 6:
            return "FF" + color.upper()
        elif len(color) == 8:
            return color.upper()
        else:
            # fallback black
            return "FF000000"

    wb = openpyxl.Workbook()
    DK = "0e1c12"   # dark green (6-digit)
    MG = "163020"   # medium green
    AG = "196b41"   # accent green
    BG_C = "30c472" # bright green
    LG = "e4f5ec"   # light green
    LG2 = "f0faf4"  # very light green
    WH = "FFFFFF"
    GR = "f5f8f5"   # gray
    TK = "0a1c12"   # text dark
    TM = "196b41"   # text medium
    AM = "c8922a"   # amber
    RD = "c0392b"   # red
    BL = "1a5ea8"   # blue
    PU = "6b2fa8"   # purple
    BD = "c0d8c8"   # border color

    # Custom functions with aRGB conversion
    def F(color, bold=False, size=10):
        return Font(name="Calibri", color=to_aRGB(color), bold=bold, size=size)

    def BF(color):
        return PatternFill("solid", fgColor=to_aRGB(color))

    def AL(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def bdr():
        s = Side(style="thin", color=to_aRGB(BD))
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_bot():
        s = Side(style="thin", color=to_aRGB(BD))
        t = Side(style="medium", color=to_aRGB(AG))
        return Border(left=s, right=s, top=s, bottom=t)

    def safe_title(t):
        c = _re.sub(r'[:/\\?*\[\]]', '', t).encode('ascii', 'ignore').decode('ascii').strip()[:31]
        return c or "Sheet"

    def xval(v):
        if v is None: return ""
        if isinstance(v, (str, int, float, bool)): return v
        if isinstance(v, dict): return v.get("name", v.get("task", v.get("story", str(v))))
        if isinstance(v, list): return "; ".join(str(xval(x)) for x in v)
        return str(v)

    def col_headers(ws, cols, row=2, bg=AG):
        for c, (lbl, w) in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=lbl)
            cell.fill = BF(bg); cell.font = F(WH, bold=True, size=10)
            cell.alignment = AL("center", "center"); cell.border = thick_bot()
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[row].height = 22

    def title_row(ws, text, n, bg=DK, fsz=14):
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        c = ws["A1"]; c.value = text; c.fill = BF(bg)
        c.font = F(WH, bold=True, size=fsz); c.alignment = AL("center", "center")
        c.border = bdr(); ws.row_dimensions[1].height = 40

    is_agile = any(k in methodology.lower() for k in ['agile','scrum','safe'])

    DISCOVERY_SPRINT = {
        "sprint_num": 0, "name": "Discovery & Setup: Foundation Sprint", "weeks": "Week 1–2",
        "goal": "Project foundation, Snowflake environment setup, RBAC, and discovery sessions",
        "sow_ref": "Standard — applies to all projects",
        "stories": [
            "Conduct Discovery Sessions with customer stakeholders",
            "Snowflake environment provisioning and configuration",
            "RBAC design and role hierarchy setup (SYSADMIN, USERADMIN, custom roles)",
            "Data source inventory and connectivity assessment",
            "Architecture Review Board (ARB) pack preparation",
            "Project Review Board (PRB) pack preparation",
            "Customer sign-off and acknowledgement of Discovery findings",
            "Sprint 1 planning and backlog grooming",
        ],
        "deliverable": "ARB/PRB packs approved · Customer acknowledgement received · Snowflake env ready",
    }
    first_name = str(wbs_sprints[0].get("name","")).lower() if wbs_sprints else ""
    if "discovery" not in first_name and "foundation" not in first_name:
        full_sprints = [DISCOVERY_SPRINT] + list(wbs_sprints)
    else:
        full_sprints = list(wbs_sprints)
        mandatory = DISCOVERY_SPRINT["stories"]
        existing = [str(xval(s)).lower() for s in full_sprints[0].get("stories", [])]
        for task in mandatory:
            if not any(task.lower()[:20] in e for e in existing):
                full_sprints[0].setdefault("stories", []).insert(0, task)

    ws1 = wb.active; ws1.title = safe_title("Project Summary")
    ws1.sheet_view.showGridLines = False; ws1.sheet_view.zoomScale = 95

    ws1.merge_cells("A1:F3"); h = ws1["A1"]; h.value = "PROJECT PLAN"
    h.fill = BF(DK); h.font = F(WH, bold=True, size=22); h.alignment = AL("left","center")
    for r in [1,2,3]: ws1.row_dimensions[r].height = 20 if r != 2 else 36

    ws1.merge_cells("A4:F4"); s = ws1["A4"]; s.value = plan_name
    s.fill = BF(MG); s.font = F(BG_C, bold=True, size=16); s.alignment = AL("left","center")
    ws1.row_dimensions[4].height = 30

    ws1.merge_cells("A5:F5"); m = ws1["A5"]
    m.value = f"Cortex Project Intel  ·  {datetime.now().strftime('%d %B %Y')}  ·  Methodology: {methodology}"
    m.fill = BF(MG); m.font = F("90d4a0", size=9); m.alignment = AL("left","center")
    ws1.row_dimensions[5].height = 18; ws1.row_dimensions[6].height = 10

    info = [("Project Name", plan_name), ("Methodology", methodology),
            ("Generated", datetime.now().strftime('%d %B %Y')),
            ("Total Sprints / Phases", str(len(full_sprints))),
            ("Discovery Phase", "Week 1–2 (Standard — all projects)"),
            ("Source", "Cortex Project Intel — Snowflake AI")]
    for i, (lbl, val) in enumerate(info, 7):
        lc = ws1.cell(row=i, column=1, value=lbl); vc = ws1.cell(row=i, column=2, value=val)
        lc.fill = BF(AG); lc.font = F(WH, bold=True, size=10); lc.alignment = AL("right","center"); lc.border = bdr()
        vc.fill = BF(LG); vc.font = F(TK, size=10); vc.alignment = AL("left","center"); vc.border = bdr()
        ws1.row_dimensions[i].height = 20

    ws1.row_dimensions[13].height = 10
    ws1.merge_cells("A14:F14"); ph = ws1["A14"]; ph.value = "FULL PLAN TEXT"
    ph.fill = BF(AG); ph.font = F(WH, bold=True, size=10); ph.alignment = AL("center","center"); ph.border = bdr()
    ws1.row_dimensions[14].height = 18
    ws1.merge_cells("A15:F90"); pc = ws1["A15"]; pc.value = plan_text
    pc.font = F(TK, size=9); pc.alignment = AL("left","top",wrap=True); pc.fill = BF(GR)
    for r in range(15, 91): ws1.row_dimensions[r].height = 13
    ws1.column_dimensions["A"].width = 22; ws1.column_dimensions["B"].width = 55
    for col in ["C","D","E","F"]: ws1.column_dimensions[col].width = 14

    ws2 = wb.create_sheet(safe_title("WBS Sprint Board"))
    ws2.sheet_view.showGridLines = False; ws2.freeze_panes = "A3"
    WBS_COLS = [("Sprint / Phase",24),("Weeks",11),("Sprint Goal",38),("Task / User Story",46),
                ("Type",14),("Story Points",12),("Priority",12),("Assignee",18),("Deliverable",34),("Status",12)]
    title_row(ws2, ("SPRINT BOARD" if is_agile else "WORK BREAKDOWN STRUCTURE") + f"  —  {plan_name}", 10)
    col_headers(ws2, WBS_COLS)

    sp_pts = [8,8,5,5,3,3,2,3]; sev_list = ["High","High","Medium","Medium","Low","Medium","Low","Medium"]
    sev_col = {"High":RD,"Medium":AM,"Low":AG}; sev_bg = {"High":"fdeaea","Medium":"fef7e8","Low":"e4f5ec"}
    row = 3
    for s_idx, sprint in enumerate(full_sprints):
        sprint_name = str(xval(sprint.get("name",""))); weeks_lbl = str(xval(sprint.get("weeks","")))
        goal = str(xval(sprint.get("goal",""))); deliverable = str(xval(sprint.get("deliverable","")))
        raw_tasks = sprint.get("stories", sprint.get("tasks",[]))
        tasks = [str(xval(t)).strip() for t in (raw_tasks if isinstance(raw_tasks, list) else []) if str(xval(t)).strip()]
        if not tasks: tasks = [goal] if goal else ["Sprint deliverable"]
        hdr_bg = "0a1a0e" if s_idx == 0 else (DK if s_idx % 2 == 0 else MG)
        hdr_fg = BG_C

        ws2.merge_cells(f"A{row}:J{row}")
        banner = ws2.cell(row=row, column=1, value=f"  {sprint_name}   |   {weeks_lbl}   |   {goal}")
        banner.fill = BF(hdr_bg); banner.font = F(hdr_fg, bold=True, size=11)
        banner.alignment = AL("left","center"); banner.border = bdr()
        ws2.row_dimensions[row].height = 26; row += 1

        for t_idx, task in enumerate(tasks):
            row_bg = "d8f0e2" if s_idx == 0 and t_idx % 2 == 0 else \
                     "c8e8d4" if s_idx == 0 else (LG if t_idx % 2 == 0 else LG2)
            pts = sp_pts[min(t_idx, len(sp_pts)-1)]
            sev = sev_list[min(t_idx, len(sev_list)-1)]
            sc = sev_col[sev]; sb = sev_bg[sev]
            t_type = "User Story" if is_agile else "Task"
            deliv = deliverable if t_idx == 0 else ""
            row_vals = [sprint_name, weeks_lbl, goal, task, t_type, pts, sev, "TBD", deliv, "To Do"]
            for c, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=row, column=c, value=val); cell.border = bdr()
                if c == 6:
                    cell.fill = BF("e3eeff"); cell.font = F(BL, bold=True); cell.alignment = AL("center","center")
                elif c == 7:
                    cell.fill = BF(sb); cell.font = F(sc, bold=True); cell.alignment = AL("center","center")
                elif c == 10:
                    cell.fill = BF("fef7e8"); cell.font = F(AM, bold=True, size=9); cell.alignment = AL("center","center")
                elif c == 9 and t_idx == 0:
                    cell.fill = BF(row_bg); cell.font = F(AG, bold=True); cell.alignment = AL("left","center",wrap=True)
                else:
                    cell.fill = BF(row_bg); cell.font = F(TK, bold=(c in [1,2] and t_idx == 0))
                    cell.alignment = AL("left","center",wrap=True)
            ws2.row_dimensions[row].height = 19; row += 1
        for c in range(1, 11):
            g = ws2.cell(row=row, column=c, value=""); g.fill = BF("d0e8d8"); g.border = bdr()
        ws2.row_dimensions[row].height = 5; row += 1

    ws3 = wb.create_sheet(safe_title("RAID Log"))
    ws3.sheet_view.showGridLines = False; ws3.freeze_panes = "A3"
    RAID_COLS = [("ID",5),("Category",15),("Type",16),("Description",46),
                 ("Likelihood",13),("Impact",12),("Mitigation",44),("Owner",20)]
    title_row(ws3, f"RAID LOG  —  {plan_name}", len(RAID_COLS))
    col_headers(ws3, RAID_COLS)

    lh_c = {"High":RD,"Medium":AM,"Low":AG}
    lh_b = {"High":"fdecea","Medium":"fef7e8","Low":"e4f5ec"}
    cat_bg = {"Risk":"fdecea","Assumption":"e3eeff","Issue":"fef7e8","Dependency":"f0e8ff"}
    cat_fg = {"Risk":RD,"Assumption":BL,"Issue":AM,"Dependency":PU}

    raid_data = []
    sec = _re.search(r'##\s*RAID\s*LOG(.*?)(?=##\s+[A-Z]|\Z)', plan_text, _re.DOTALL|_re.IGNORECASE)
    if sec:
        skip = {'','id','type','category','description','likelihood','impact','mitigation','owner','---'}
        for line in sec.group(1).split('\n'):
            parts = [p.strip() for p in line.split('|') if p.strip()]
            if len(parts) >= 4 and not all(set(p) <= set('-: ') for p in parts):
                if parts[0].lower() not in skip:
                    raid_data.append(parts)

    discovery_raid = [
        ["","Risk","Technical","Snowflake environment not provisioned before Week 1","High","High","Pre-provision environment Day 1; track in daily standup","Platform Engineer"],
        ["","Risk","Process","ARB/PRB approval delayed beyond Week 2","High","High","Submit ARB/PRB packs by end of Day 3; escalate if no response by Day 8","Project Manager"],
        ["","Assumption","Process","Customer stakeholders available for Discovery Sessions in Week 1–2","—","—","Confirm attendance list before project kick-off","Project Manager"],
        ["","Dependency","External","Customer sign-off on Discovery findings required before Sprint 2","High","High","Schedule sign-off meeting end of Week 2","Project Manager"],
        ["","Risk","Resource","RBAC roles not approved by InfoSec in time","Medium","High","Raise InfoSec request Week 1 Day 1; track daily","Platform Engineer"],
    ]

    if not raid_data:
        raid_data = [
            ["","Risk","Technical","Data access permissions not provisioned","High","High","Pre-provision all Snowflake roles before Sprint 1","Platform Engineer"],
            ["","Risk","Timeline","Scope creep from stakeholder change requests","Medium","High","Freeze scope after Sprint 2; route changes via CR process","Project Manager"],
            ["","Risk","Resource","Key person dependency on lead architect","Medium","Medium","Cross-train 2 team members on architecture decisions","Tech Lead"],
            ["","Risk","Budget","Compute cost overrun on large warehouse queries","Low","High","Set resource monitor alerts at 80% of credit budget","FinOps"],
            ["","Assumption","Process","Team availability maintained at 80% throughout","—","—","Confirm resource allocation with line managers before start","PMO"],
            ["","Assumption","Technical","Source APIs remain stable during integration","—","—","Document API versions; add version-check step","Data Engineer"],
            ["","Dependency","External","InfoSec approval for external API access","High","High","Raise InfoSec request Week 1; track weekly","Project Manager"],
            ["","Dependency","Data","Source data quality meets >90% DQ threshold","Medium","High","Run DQ profiling Sprint 1; escalate if score <90%","Data Engineer"],
            ["","Issue","Process","Legacy data quality issues in source tables","—","Medium","Run cleansing pipeline in Sprint 1; document exceptions","Data Engineer"],
            ["","Issue","Technical","Snowflake credit limitations on trial account","Low","Low","Monitor credit usage daily; upgrade account if needed","Platform Engineer"],
        ]

    all_raid = discovery_raid + raid_data

    for i, r in enumerate(all_raid[:30]):
        while len(r) < 8: r.append("")
        rn = i + 3
        cat = str(r[1]).strip().title(); lh = str(r[4]).strip().title(); im = str(r[5]).strip().title()
        rbg = "d0e8d8" if i < len(discovery_raid) else (LG if i % 2 == 0 else LG2)
        for c, val in enumerate(r[:8], 1):
            cell = ws3.cell(row=rn, column=c, value=str(xval(val))); cell.border = bdr()
            if c == 1:
                cell.value = i + 1; cell.fill = BF(AG); cell.font = F(WH, bold=True); cell.alignment = AL("center","center")
            elif c == 2:
                cell.fill = BF(cat_bg.get(cat,"f5f8f5")); cell.font = F(cat_fg.get(cat,TK), bold=True); cell.alignment = AL("center","center")
            elif c == 3:
                cell.fill = BF("f0f0ff"); cell.font = F(BL); cell.alignment = AL("center","center")
            elif c == 5:
                bg = lh_b.get(lh,"f5f8f5") if lh not in ("—","") else GR
                cell.fill = BF(bg); cell.font = F(lh_c.get(lh,TK), bold=True); cell.alignment = AL("center","center")
            elif c == 6:
                bg = lh_b.get(im,"f5f8f5") if im not in ("—","") else GR
                cell.fill = BF(bg); cell.font = F(lh_c.get(im,TK), bold=True); cell.alignment = AL("center","center")
            else:
                cell.fill = BF(rbg); cell.font = F(TK); cell.alignment = AL("left","center",wrap=True)
        ws3.row_dimensions[rn].height = 22

    ws3.insert_rows(3)
    ws3.merge_cells(f"A3:{get_column_letter(len(RAID_COLS))}3")
    note = ws3["A3"]; note.value = "  ★  DISCOVERY PHASE RISKS (Week 1–2 — Standard for all projects)  ★"
    note.fill = BF("0a2a14"); note.font = F(BG_C, bold=True, size=10); note.alignment = AL("center","center")
    ws3.row_dimensions[3].height = 18

    ws4 = wb.create_sheet(safe_title("Test Cases"))
    ws4.sheet_view.showGridLines = False; ws4.freeze_panes = "A3"
    TC_COLS = [("ID",8),("Priority",12),("Test Type",16),("Scenario",34),
               ("Preconditions",28),("Test Steps",46),("Expected Result",36),
               ("Actual Result",36),("Status",13)]
    title_row(ws4, f"TEST CASES  —  {plan_name}", len(TC_COLS))
    col_headers(ws4, TC_COLS)

    pri_c = {"High":RD,"Medium":AM,"Low":AG}
    pri_b = {"High":"fdecea","Medium":"fef7e8","Low":"e4f5ec"}
    typ_b = {"Integration":"e3eeff","Data Quality":"e4f5ec","Security":"fdecea",
             "Performance":"fef7e8","Functional":"f0e8ff","Regression":"f5f5f5","UAT":"e4f5ec"}
    typ_f = {"Integration":BL,"Data Quality":AG,"Security":RD,
             "Performance":AM,"Functional":PU,"Regression":"555555","UAT":AG}

    tc_data = []
    tcs = _re.search(r'##\s*TEST\s*CASES(.*?)(?=##\s+[A-Z]|\Z)', plan_text, _re.DOTALL|_re.IGNORECASE)
    if tcs:
        skip2 = {'','id','scenario','steps','expected result','pass criteria','priority','test type','type','---'}
        for line in tcs.group(1).split('\n'):
            parts = [p.strip() for p in line.split('|') if p.strip()]
            if len(parts) >= 4 and not all(set(p) <= set('-: ') for p in parts):
                if parts[0].lower() not in skip2:
                    tc_data.append(parts)

    discovery_tc = [
        ["DTC-01","High","Functional","Snowflake environment connectivity","Account provisioned","1.Login to Snowflake  2.Run SELECT CURRENT_VERSION()","Version returned; no errors","","Not Started"],
        ["DTC-02","High","Security","RBAC roles correctly provisioned","Roles created per design","1.Login as each role  2.Test permitted objects  3.Test restricted objects","Permitted access granted; restricted access denied","","Not Started"],
        ["DTC-03","High","Functional","ARB/PRB pack completeness review","Discovery sessions complete","1.Review ARB pack checklist  2.Review PRB pack checklist  3.Confirm customer sign-off","All sections complete; customer signed off","","Not Started"],
        ["DTC-04","Medium","Functional","Discovery session outputs documented","Sessions conducted","1.Review meeting notes  2.Verify all data sources captured  3.Confirm scope agreed","All outputs documented; scope baseline confirmed","","Not Started"],
    ]

    if not tc_data:
        tc_data = [
            ["TC-01","High","Integration","End-to-end pipeline ingestion","Source data available; pipeline deployed","1. Stage source CSV  2. Execute pipeline  3. Query target table","Row count matches source; no nulls in key columns","","Not Started"],
            ["TC-02","High","Data Quality","Schema and column type validation","Target table created","1. Run schema check SQL  2. Validate types  3. Check PK uniqueness","Zero schema violations; zero duplicate PKs","","Not Started"],
            ["TC-03","High","Security","Role-based access control","VIEWER and EDITOR roles provisioned","1. Login as VIEWER  2. Attempt INSERT  3. Verify permission error","INSERT denied; SELECT succeeds for VIEWER","","Not Started"],
            ["TC-04","High","Functional","Cortex AI response validation","CORTEX_COMPLETE enabled","1. Call CORTEX_COMPLETE  2. Parse JSON  3. Validate schema","Valid JSON; all fields present; latency < 5s","","Not Started"],
            ["TC-05","Medium","Performance","Dashboard load time","App deployed; tables populated","1. Open app URL  2. Load dashboard  3. Measure render time","Dashboard interactive within 3 seconds","","Not Started"],
            ["TC-06","Medium","Data Quality","DQ rule execution","DQ rule set configured","1. Execute DQ rules  2. Check DQ score  3. Review error log","DQ score >= 95%; flagged rows in error log","","Not Started"],
            ["TC-07","Medium","Regression","Pipeline idempotency","Pipeline run at least once","1. Re-trigger pipeline  2. Query row count  3. Compare pre/post","Row count unchanged; zero duplicates","","Not Started"],
            ["TC-08","Low","UAT","Business stakeholder sign-off","All TC-01–07 passed","1. Demo to business owner  2. Walk through KPIs  3. Capture sign-off","Business owner approves; sign-off document saved","","Not Started"],
        ]

    all_tc = discovery_tc + tc_data
    for i, rd in enumerate(all_tc[:25]):
        while len(rd) < 9: rd.append("")
        rn = i + 3; rbg = "d0e8d8" if i < len(discovery_tc) else (LG if i % 2 == 0 else LG2)
        pri = str(rd[1]).strip().title(); tt = str(rd[2]).strip().title()
        for c, val in enumerate(rd[:9], 1):
            cell = ws4.cell(row=rn, column=c, value=str(xval(val))); cell.border = bdr()
            if c == 1:
                cell.fill = BF(AG if i >= len(discovery_tc) else "0a2a14")
                cell.font = F(WH, bold=True); cell.alignment = AL("center","center")
            elif c == 2:
                cell.fill = BF(pri_b.get(pri,"fef7e8")); cell.font = F(pri_c.get(pri,AM), bold=True); cell.alignment = AL("center","center")
            elif c == 3:
                cell.fill = BF(typ_b.get(tt,"f5f5f5")); cell.font = F(typ_f.get(tt,TK), bold=True); cell.alignment = AL("center","center")
            elif c == 8:
                cell.fill = BF("fafcfb"); cell.font = F("bbbbbb", size=9); cell.alignment = AL("left","top",wrap=True)
                if not cell.value: cell.value = "— fill after execution —"
            elif c == 9:
                cell.fill = BF("fef7e8"); cell.font = F(AM, bold=True); cell.alignment = AL("center","center")
            else:
                cell.fill = BF(rbg); cell.font = F(TK); cell.alignment = AL("left","top",wrap=True)
        ws4.row_dimensions[rn].height = 44

    ws5 = wb.create_sheet(safe_title("Delivery Timeline"))
    ws5.sheet_view.showGridLines = False; ws5.freeze_panes = "A3"
    TL_COLS = [("Week / Period",15),("Sprint / Phase",28),("Milestone",50),
               ("Tasks",44),("Owner",20),("Status",14),("Notes",28)]
    title_row(ws5, f"DELIVERY TIMELINE  —  {plan_name}", len(TL_COLS))
    col_headers(ws5, TL_COLS)

    row = 3
    for s_idx, sprint in enumerate(full_sprints):
        sprint_name = str(xval(sprint.get("name",""))); weeks_lbl = str(xval(sprint.get("weeks","")))
        deliverable = str(xval(sprint.get("deliverable",""))); goal = str(xval(sprint.get("goal","")))
        raw_tasks = sprint.get("stories", sprint.get("tasks",[]))
        tasks = [str(xval(t)).strip() for t in (raw_tasks if isinstance(raw_tasks, list) else []) if str(xval(t)).strip()]
        if not tasks: tasks = [goal] if goal else ["Deliver sprint objectives"]

        is_disc = s_idx == 0 and ("discovery" in sprint_name.lower() or "foundation" in sprint_name.lower())
        hdr_bg = "0a1a0e" if is_disc else (DK if s_idx % 2 == 0 else MG)
        rbg    = "c8e8d0" if is_disc else (LG if s_idx % 2 == 0 else LG2)

        hdr = [weeks_lbl, sprint_name, f"Deliverable: {deliverable}", goal, "Project Team","Planned",""]
        for c, val in enumerate(hdr, 1):
            cell = ws5.cell(row=row, column=c, value=str(xval(val))); cell.border = bdr()
            if c <= 2:
                cell.fill = BF(hdr_bg); cell.font = F(BG_C, bold=True, size=11); cell.alignment = AL("left","center")
            elif c == 3:
                cell.fill = BF(rbg); cell.font = F(AG, bold=True); cell.alignment = AL("left","center",wrap=True)
            elif c == 6:
                cell.fill = BF("fef7e8"); cell.font = F(AM, bold=True); cell.alignment = AL("center","center")
            else:
                cell.fill = BF(rbg); cell.font = F(TK); cell.alignment = AL("left","center",wrap=True)
        ws5.row_dimensions[row].height = 26; row += 1

        for t_idx, task in enumerate(tasks):
            tbg = "e8f5ee" if is_disc else ("f7faf8" if s_idx%2==0 else "f2f7f3")
            tv = ["","", f"  • {task}","","TBD","To Do",""]
            for c, val in enumerate(tv, 1):
                cell = ws5.cell(row=row, column=c, value=str(xval(val))); cell.border = bdr()
                if c == 3:
                    cell.fill = BF(tbg); cell.font = F(TM, size=10); cell.alignment = AL("left","center",wrap=True)
                elif c == 6:
                    cell.fill = BF("fef7e8"); cell.font = F(AM, bold=True, size=9); cell.alignment = AL("center","center")
                else:
                    cell.fill = BF(tbg); cell.font = F(TK, size=9); cell.alignment = AL("left","center")
            ws5.row_dimensions[row].height = 17; row += 1

        for c in range(1, 8):
            g = ws5.cell(row=row, column=c, value=""); g.fill = BF("c8e0d0"); g.border = bdr()
        ws5.row_dimensions[row].height = 4; row += 1

    # Stakeholder Matrix Sheet (hardcoded, can be replaced with generated map if desired)
    ws6 = wb.create_sheet(safe_title("Stakeholder Matrix"))
    ws6.sheet_view.showGridLines = False
    SM_COLS = [("Stakeholder",28),("Role",22),("Interest Level",16),("Influence",16),
               ("Communication Freq",20),("Preferred Channel",22),("Key Concerns",40),("Engagement Strategy",40)]
    title_row(ws6, f"STAKEHOLDER MATRIX  —  {plan_name}", len(SM_COLS))
    col_headers(ws6, SM_COLS)
    stakeholders = [
        ["Project Sponsor","Executive","High","High","Weekly","Email/Steering","Budget & ROI delivery","Executive steering committee; weekly status"],
        ["Business Owner","Senior Manager","High","High","Weekly","Dashboard","Scope & timeline","Weekly demo; sign-off gates"],
        ["Data Architect","Technical","High","Medium","Daily","Standup","Architecture decisions","Daily standup; architecture reviews"],
        ["Data Engineers","Technical","Medium","Medium","Daily","Standup","Task clarity","Sprint planning; daily standups"],
        ["QA Lead","Technical","High","Medium","Weekly","Jira/Email","Test coverage","UAT sign-off gates; defect triage"],
        ["InfoSec","Governance","Medium","High","Fortnightly","Email","Security & compliance","Fortnightly check-ins; RBAC approvals"],
        ["End Users","Business","Medium","Low","Monthly","Training","Ease of use","UAT sessions; training plan"],
        ["PMO","Governance","Low","Medium","Weekly","Report","Governance & reporting","Weekly status report; risk escalation"],
    ]
    for i, row_data in enumerate(stakeholders):
        rn = i + 3; rbg = LG if i % 2 == 0 else LG2
        int_col = {"High":"#196b41","Medium":"#c8922a","Low":"#888888"}
        int_bg  = {"High":"e4f5ec","Medium":"fef7e8","Low":"f5f5f5"}
        for c, val in enumerate(row_data, 1):
            cell = ws6.cell(row=rn, column=c, value=val); cell.border = bdr()
            if c == 1:
                cell.fill = BF(AG); cell.font = F(WH, bold=True); cell.alignment = AL("left","center")
            elif c in (3, 4):
                lv = val.strip().title()
                cell.fill = BF(int_bg.get(lv, "f5f5f5"))
                cell.font = F(int_col.get(lv, TK), bold=True); cell.alignment = AL("center","center")
            else:
                cell.fill = BF(rbg); cell.font = F(TK); cell.alignment = AL("left","center",wrap=True)
        ws6.row_dimensions[rn].height = 22

    output = io.BytesIO(); wb.save(output); return output.getvalue()


if st.session_state.past_projects_df is None:
    st.session_state.past_projects_df = load_past_projects()
past_projects_df = st.session_state.past_projects_df
db_names_set = set(past_projects_df['project_name'].tolist())

with st.sidebar:
    st.markdown("""
    <div style="background:linear-gradient(160deg,#0d3d22 0%,#196b41 60%,#228b55 100%);
                padding:1.2rem 1.1rem 1rem;margin-bottom:0.75rem;position:relative;overflow:hidden;">
        <div style="position:absolute;top:-20px;right:-20px;width:80px;height:80px;
                    background:rgba(48,196,114,0.08);border-radius:50%;"></div>
        <div style="display:flex;align-items:center;gap:0.7rem;margin-bottom:0.5rem;">
            <div style="width:36px;height:36px;background:rgba(255,255,255,0.13);border-radius:10px;
                        display:flex;align-items:center;justify-content:center;font-size:1rem;
                        border:1px solid rgba(255,255,255,0.20);">📊</div>
            <div>
                <div style="font-size:9px;font-weight:700;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.14em;font-family:'Syne',sans-serif;">Snowflake Cortex</div>
                <div style="font-size:15px;font-weight:800;color:#fff;letter-spacing:-0.01em;line-height:1.2;font-family:'Syne',sans-serif;">Project Intel</div>
            </div>
        </div>
        <div style="font-size:10px;color:rgba(255,255,255,0.38);line-height:1.55;">
            Configure your project. Cortex selects the best templates automatically.
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### 📄 Statement of Work")
    sow_pdf  = st.file_uploader("Upload SOW PDF (mandatory)", type=["pdf"])
    sow_text = ""
    if sow_pdf is not None:
        if PDF_AVAILABLE:
            try:
                reader   = PyPDF2.PdfReader(sow_pdf)
                sow_text = "\n".join([p.extract_text() or "" for p in reader.pages])
                st.success("✅ SOW loaded ({} chars)".format(len(sow_text)))
            except Exception as e:
                st.error(f"PDF read error: {e}")
        else:
            st.warning("⚠️ PyPDF2 not installed.")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### 📂 Project Plans (CSV)")
    plans_csvs = st.file_uploader("Upload CSVs (multiple allowed)", type=["csv"], accept_multiple_files=True)
    existing_names = {p["project_name"] for p in st.session_state.csv_projects}
    if plans_csvs:
        newly_added = 0
        for f in plans_csvs:
            proj = parse_csv_as_project(f)
            if proj and proj["project_name"] not in existing_names:
                st.session_state.csv_projects.append(proj)
                existing_names.add(proj["project_name"])
                newly_added += 1
        if newly_added:
            st.success("✅ {} new plan(s) added — {} total.".format(newly_added, len(st.session_state.csv_projects)))
        else:
            st.info("{} plan(s) already loaded.".format(len(st.session_state.csv_projects)))
    if st.session_state.csv_projects:
        st.markdown("<div style='font-size:11px;color:var(--green-mid);font-weight:700;margin:.35rem 0;'>📎 {} plan(s) in memory</div>".format(len(st.session_state.csv_projects)), unsafe_allow_html=True)
        for idx_p, proj in enumerate(st.session_state.csv_projects):
            col_a, col_b = st.columns([3, 1])
            with col_a:
                nm = proj['project_name']
                already = nm in db_names_set or nm in st.session_state.imported_csv_names
                badge = "✅" if already else "⏳"
                name_color = "var(--green-mid)" if already else "var(--gold)"
                st.markdown("<div style='font-size:11px;color:var(--text-secondary);padding:.2rem 0;'>📄 {} <span style='color:{};font-size:9px;'>{}</span></div>".format(proj['project_name'], name_color, badge), unsafe_allow_html=True)
            with col_b:
                btn_label = "Re-save" if (proj['project_name'] in db_names_set or proj['project_name'] in st.session_state.imported_csv_names) else "Import"
                if st.button(btn_label, key="import_csv_{}".format(idx_p)):
                    try:
                        pn=proj["project_name"].replace("'","''")
                        pd_=proj["description"][:500].replace("'","''")
                        wb_=proj["wbs_summary"][:3000].replace("'","''")
                        sd=f"'{proj['start_date']}'" if proj["start_date"] else "NULL"
                        ed=f"'{proj['end_date']}'" if proj["end_date"] else "NULL"
                        try: session.sql(f"DELETE FROM projects WHERE project_name='{pn}' AND description LIKE '%CSV%'").collect()
                        except: pass
                        session.sql(f"INSERT INTO projects (project_name,description,start_date,end_date,status,lead_architect,technologies_used,wbs_summary,risk_log_summary,test_cases_summary,deployment_plan) VALUES('{pn}','{pd_}',{sd},{ed},'Completed','','','{wb_}','','','')").collect()
                        st.session_state.imported_csv_names.add(proj["project_name"])
                        st.session_state.past_projects_df = load_past_projects()
                        st.success("✅ Saved '{}'".format(proj['project_name']))
                        st.rerun()
                    except Exception as e:
                        st.error(f"Import failed: {e}")
        if st.button("🗑️ Clear All CSV Plans", key="clear_csvs"):
            st.session_state.csv_projects = []; st.session_state.imported_csv_names = set(); st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### 🆕 New Project Details")
    new_project_name = st.text_input("Project Name", "Data Cataloging Initiative")
    new_project_desc = st.text_area("Description", "A new data cataloging initiative to index all data assets in Snowflake.", height=80)
    new_project_tech = st.text_input("Key Technologies", "Snowflake, Streamlit, Python")
    customer_name    = st.text_input("Customer / Organisation", "Acme Corp")
    project_type     = st.selectbox("Project Type", ["Data Engineering","Analytics / BI","Data Science / ML","Data Platform Migration","App Development","Data Governance","Other"])
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### ⚙️ Plan Settings")
    team_size      = st.slider("Team Size", 2, 20, 5)
    duration_weeks = st.slider("Target Duration (weeks)", 2, 52, 12)
    risk_appetite  = st.select_slider("Risk Appetite", ["Low","Medium","High"], value="Medium")
    methodology    = st.selectbox("Methodology", ["Agile / Scrum","Waterfall","Hybrid","SAFe"])
    delivery_priority = st.selectbox("Delivery Priority", ["Balanced (scope, time, cost)","Time-boxed (fixed deadline)","Scope-driven (all requirements)","Cost-constrained (fixed budget)"])
    team_experience   = st.selectbox("Team Experience", ["Expert (3+ yrs)","Intermediate (1-3 yrs)","Beginner (< 1 yr)","Mixed levels"])
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### 👥 Team Roles")
    rc1, rc2 = st.columns(2)
    with rc1:
        role_pm = st.number_input("👔 PM",  min_value=0, max_value=5, value=1)
        role_sa = st.number_input("🏗️ SA", min_value=0, max_value=5, value=1)
        role_de = st.number_input("⚙️ DE", min_value=0, max_value=10, value=2)
    with rc2:
        role_ds = st.number_input("🔬 DS",  min_value=0, max_value=5, value=1)
        role_qa = st.number_input("🧪 QA", min_value=0, max_value=5, value=1)
    has_dedicated_pm = role_pm > 0; has_qa_resource = role_qa > 0
    team_roster = "PM×{}, SA×{}, DE×{}, DS×{}, QA×{}".format(role_pm, role_sa, role_de, role_ds, role_qa)
    team_size_from_roles = role_pm + role_sa + role_de + role_ds + role_qa
    if team_size < team_size_from_roles: team_size = team_size_from_roles
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### 🔒 Compliance & Data")
    compliance_reqs  = st.multiselect("Compliance", ["GDPR","HIPAA","SOC 2","PCI-DSS","ISO 27001","None"], default=["None"])
    has_data_sensitivity = st.selectbox("Data Sensitivity", ["Public","Internal","Confidential","Restricted / PII"])
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='sb-card'>", unsafe_allow_html=True)
    st.markdown("### 💰 ROI Inputs")
    manual_hours = st.number_input("Manual Planning Hours (PM+SA)", min_value=1, max_value=100, value=12, help="Total hours for PM+SA to create project plan manually")
    cortex_hours = st.number_input("Cortex Planning Hours (PM+SA)", min_value=0, max_value=20, value=4, help="Hours with Cortex-assisted planning")
    risks_identified = st.slider("Risks Identified (total per project)", 0, 15, 5, help="Number of risks proactively identified and mitigated")
    hourly_rate = 80
    risk_saving_per_week_per_risk = 500
    risk_saving = risks_identified * duration_weeks * risk_saving_per_week_per_risk
    planning_saving = (manual_hours - cortex_hours) * hourly_rate
    total_roi_value = planning_saving + risk_saving
    ai_planning_cost = cortex_hours * hourly_rate
    net_benefit = total_roi_value - ai_planning_cost
    roi_percentage = int((net_benefit / max(ai_planning_cost, 1)) * 100)
    time_saved_pct = int(((manual_hours - cortex_hours) / max(manual_hours, 1)) * 100)
    payback_days = round(ai_planning_cost / max(planning_saving + risk_saving, 1) * manual_hours, 1)
    st.session_state.roi_data = {
        "hourly_rate": hourly_rate,
        "manual_hours": manual_hours,
        "cortex_hours": cortex_hours,
        "risks_identified": risks_identified,
        "planning_saving": planning_saving,
        "risk_saving": risk_saving,
        "total_roi_value": total_roi_value,
        "ai_planning_cost": ai_planning_cost,
        "net_benefit": net_benefit,
        "roi_percentage": roi_percentage,
        "time_saved_pct": time_saved_pct,
        "payback_days": payback_days,
        "duration_weeks": duration_weeks,
        "risk_saving_per_week_per_risk": risk_saving_per_week_per_risk
    }
    st.markdown("</div>", unsafe_allow_html=True)

    timeline_score = max(10, min(95, 100 - max(0, duration_weeks - 16) * 2))
    team_score     = max(20, min(95, 60 + (team_size_from_roles * 4)))
    risk_score     = {"Low": 90, "Medium": 68, "High": 42}.get(risk_appetite, 68)
    compliance_penalty = 8 * max(0, len([c for c in compliance_reqs if c != "None"]) - 1)
    risk_score = max(20, risk_score - compliance_penalty)
    experience_score = {"Expert (3+ yrs)":88,"Intermediate (1-3 yrs)":70,"Beginner (< 1 yr)":45,"Mixed levels":62}.get(team_experience, 70)

    def score_color(s):
        if s >= 75: return "#145c32", "#196b41", "rgba(25,107,65,0.4)"
        if s >= 50: return "#8a5a00", "#c8922a", "rgba(200,146,42,0.4)"
        return "#9b1c1c", "#c0392b", "rgba(192,57,43,0.4)"

    bars = [("Timeline Health", timeline_score), ("Team Capacity", team_score),
            ("Risk Posture", risk_score), ("Team Readiness", experience_score)]

    bars_html = ""
    for label, score in bars:
        bg, fg, glow = score_color(score)
        bars_html += (
            "<div style='margin-bottom:0.45rem;'>"
            "<div style='display:flex;justify-content:space-between;font-size:10px;font-weight:600;"
            "color:var(--text-secondary);margin-bottom:0.18rem;'>"
            "<span>" + label + "</span>"
            "<span style='color:" + fg + ";font-weight:700;'>" + str(score) + "%</span></div>"
            "<div style='height:5px;background:rgba(0,0,0,0.07);border-radius:999px;overflow:hidden;'>"
            "<div style='height:100%;width:" + str(score) + "%;border-radius:999px;"
            "background:linear-gradient(90deg," + bg + "," + fg + ");'></div>"
            "</div></div>"
        )

    overall = int((timeline_score + team_score + risk_score + experience_score) / 4)
    ov_bg, ov_fg, _ = score_color(overall)
    ov_label = "Healthy" if overall >= 75 else "Moderate" if overall >= 50 else "At Risk"

    health_html = (
        "<div style='background:#ffffff;border:1.5px solid var(--border);"
        "border-radius:12px;padding:0.85rem 1rem;margin-bottom:0.75rem;box-shadow:var(--shadow-xs);'>"
        "<div style='display:flex;align-items:center;gap:0.35rem;margin-bottom:0.6rem;'>"
        "<span style='font-size:0.9rem;'>📈</span>"
        "<span style='font-size:9px;font-weight:800;color:var(--green-mid);text-transform:uppercase;letter-spacing:0.18em;font-family:Syne,sans-serif;'>Project Health Score</span>"
        "<span style='margin-left:auto;background:linear-gradient(135deg," + ov_bg + "," + ov_fg + ");"
        "color:#fff;font-size:10px;font-weight:800;padding:0.15rem 0.6rem;border-radius:12px;"
        "font-family:Syne,sans-serif;'>"
        + str(overall) + "% — " + ov_label + "</span></div>"
        + bars_html +
        "</div>"
    )
    st.markdown(health_html, unsafe_allow_html=True)

    story_points_per_sprint = {"Expert (3+ yrs)": 42, "Intermediate (1-3 yrs)": 30,
                                "Beginner (< 1 yr)": 18, "Mixed levels": 26}.get(team_experience, 30)
    num_sprints_est = max(1, duration_weeks // 2)
    total_sp        = story_points_per_sprint * num_sprints_est
    dev_days        = role_de * duration_weeks * 4
    qa_days         = role_qa * duration_weeks * 3
    pm_overhead_pct = 15 if role_pm > 0 else 8

    st.markdown("""
    <div style="background:var(--green-xlight);border:1.5px solid var(--border-strong);border-radius:12px;padding:0.85rem 1rem;margin-bottom:0.75rem;box-shadow:var(--shadow-xs);">
        <div style="font-size:9px;font-weight:800;color:var(--green-mid);text-transform:uppercase;letter-spacing:0.18em;margin-bottom:0.6rem;font-family:Syne,sans-serif;">⚡ Effort Estimator</div>
        <div style='display:flex;justify-content:space-between;align-items:center;padding:0.25rem 0;border-bottom:1px solid var(--border);font-size:11px;'><span style='color:var(--text-muted);font-weight:500;'>Total Sprints</span><span style='color:var(--green-deep);font-weight:700;'>{} × 2-week</span></div>
        <div style='display:flex;justify-content:space-between;align-items:center;padding:0.25rem 0;border-bottom:1px solid var(--border);font-size:11px;'><span style='color:var(--text-muted);font-weight:500;'>Story Points</span><span style='color:var(--green-deep);font-weight:700;'>~{} SP total</span></div>
        <div style='display:flex;justify-content:space-between;align-items:center;padding:0.25rem 0;border-bottom:1px solid var(--border);font-size:11px;'><span style='color:var(--text-muted);font-weight:500;'>Dev Effort</span><span style='color:var(--green-deep);font-weight:700;'>{} person-days</span></div>
        <div style='display:flex;justify-content:space-between;align-items:center;padding:0.25rem 0;border-bottom:1px solid var(--border);font-size:11px;'><span style='color:var(--text-muted);font-weight:500;'>QA Effort</span><span style='color:var(--green-deep);font-weight:700;'>{} person-days</span></div>
        <div style='display:flex;justify-content:space-between;align-items:center;padding:0.25rem 0;border-bottom:1px solid var(--border);font-size:11px;'><span style='color:var(--text-muted);font-weight:500;'>PM Overhead</span><span style='color:var(--green-deep);font-weight:700;'>{}% of total</span></div>
        <div style='display:flex;justify-content:space-between;align-items:center;padding:0.25rem 0;font-size:11px;'><span style='color:var(--text-muted);font-weight:500;'>Team Size</span><span style='color:var(--green-deep);font-weight:700;'>{} people</span></div>
    </div>
    """.format(num_sprints_est, total_sp, dev_days, qa_days, pm_overhead_pct, team_size_from_roles), unsafe_allow_html=True)

    tips = {
        "Agile / Scrum": ("🤖 Agile Tip", "For Snowflake projects, lock your schema contract in Sprint 1. Changing column types mid-sprint adds hidden rework — define a data contract first."),
        "Waterfall":     ("📐 Waterfall Tip", "Allocate 25-30% of your timeline to UAT. Waterfall projects often underestimate the time needed for sign-off cycles with stakeholders."),
        "Hybrid":        ("🔀 Hybrid Tip", "Use Agile for data pipeline work and Waterfall gates for stakeholder milestones. This balances delivery speed with governance requirements."),
        "SAFe":          ("🏢 SAFe Tip", "Define your Solution Train dependencies in PI Planning Week 1. Cross-team blockers are the #1 SAFe delivery risk."),
    }
    tip_title, tip_text = tips.get(methodology, tips["Agile / Scrum"])
    tip_html = (
        "<div style='background:var(--gold-bg);border:1.5px solid rgba(200,146,42,0.22);"
        "border-left:4px solid var(--gold);border-radius:0 12px 12px 0;"
        "padding:0.8rem 0.9rem;margin-bottom:0.75rem;'>"
        "<div style='font-size:9px;font-weight:800;color:var(--gold);text-transform:uppercase;"
        "letter-spacing:0.16em;margin-bottom:0.3rem;font-family:Syne,sans-serif;'>✨ " + tip_title + "</div>"
        "<div style='font-size:11px;color:var(--text-primary);line-height:1.65;font-weight:500;'>" + tip_text + "</div>"
        "</div>"
    )
    st.markdown(tip_html, unsafe_allow_html=True)

    method_info = {
        "Agile / Scrum": {"icon":"🏃","color":"var(--green-mid)","cadence":"2-week sprints","ceremonies":"Daily standup, Sprint Review, Retro","best_for":"Iterative data products"},
        "Waterfall":     {"icon":"📊","color":"var(--blue)","cadence":"Sequential phases","ceremonies":"Phase gate reviews","best_for":"Fixed-scope migrations"},
        "Hybrid":        {"icon":"🔀","color":"var(--purple)","cadence":"Agile dev + Waterfall gates","ceremonies":"Sprint reviews + phase sign-offs","best_for":"Enterprise data programs"},
        "SAFe":          {"icon":"🏢","color":"var(--gold)","cadence":"PI (8-12 weeks)","ceremonies":"PI Planning, ART Sync","best_for":"Large multi-team programs"},
    }
    mi = method_info.get(methodology, method_info["Agile / Scrum"])
    method_color = mi['color']
    st.markdown(
        "<div style='background:#ffffff;border:1.5px solid var(--border);"
        "border-left:3px solid " + method_color + ";border-radius:12px;"
        "padding:0.9rem 1rem;margin-bottom:0.75rem;box-shadow:var(--shadow-xs);'>"
        "<div style='display:flex;align-items:center;gap:0.5rem;margin-bottom:0.5rem;'>"
        "<span style='font-size:1.1rem;'>" + mi['icon'] + "</span>"
        "<span style='font-size:10px;font-weight:800;color:" + method_color + ";text-transform:uppercase;letter-spacing:0.14em;font-family:Syne,sans-serif;'>" + methodology + "</span>"
        "</div>"
        "<div style='display:flex;flex-direction:column;gap:0.3rem;'>"
        "<div style='font-size:10px;color:var(--text-secondary);'><span style='font-weight:700;color:var(--text-primary);'>Cadence:</span> " + mi['cadence'] + "</div>"
        "<div style='font-size:10px;color:var(--text-secondary);'><span style='font-weight:700;color:var(--text-primary);'>Ceremonies:</span> " + mi['ceremonies'] + "</div>"
        "<div style='font-size:10px;color:var(--text-secondary);'><span style='font-weight:700;color:var(--text-primary);'>Best for:</span> " + mi['best_for'] + "</div>"
        "</div></div>",
        unsafe_allow_html=True
    )

    vel_per_sprint = story_points_per_sprint
    total_projected = vel_per_sprint * num_sprints_est
    velocity_rows = ""
    for i in range(min(num_sprints_est, 6)):
        fill = min(100, int((i+1) / num_sprints_est * 100))
        velocity_rows += (
            f"<div style='display:flex;align-items:center;gap:0.5rem;margin-bottom:0.3rem;'>"
            f"<span style='font-size:9px;color:var(--text-muted);min-width:48px;font-weight:600;'>S{i+1}</span>"
            f"<div style='flex:1;height:6px;background:var(--border);border-radius:999px;overflow:hidden;'>"
            f"<div style='height:100%;width:{fill}%;background:linear-gradient(90deg,var(--green-mid),var(--green-bright));border-radius:999px;'></div>"
            f"</div>"
            f"<span style='font-size:9px;color:var(--green-mid);font-weight:700;min-width:28px;text-align:right;'>{vel_per_sprint}sp</span>"
            f"</div>"
        )
    st.markdown(
        "<div style='background:#ffffff;border:1.5px solid var(--border);border-radius:12px;padding:0.85rem 1rem;margin-bottom:0.75rem;box-shadow:var(--shadow-xs);'>"
        "<div style='font-size:9px;font-weight:800;color:var(--green-mid);text-transform:uppercase;letter-spacing:0.18em;margin-bottom:0.55rem;font-family:Syne,sans-serif;'>🚀 Velocity Forecast</div>"
        + velocity_rows +
        f"<div style='font-size:10px;color:var(--text-muted);margin-top:0.4rem;border-top:1px solid var(--border);padding-top:0.35rem;'>"
        f"Total projected: <strong style='color:var(--green-mid);'>{total_projected} story points</strong></div>"
        "</div>",
        unsafe_allow_html=True
    )

    st.markdown("---")
    generate_button = st.button("🚀 Generate Project Plan", type="primary")


roi_data = st.session_state.get("roi_data", {
    "hourly_rate": 80,
    "manual_hours": 12,
    "cortex_hours": 4,
    "risks_identified": 5,
    "duration_weeks": 12,
    "planning_saving": (12-4)*80,
    "risk_saving": 5*12*500,
    "total_roi_value": (12-4)*80 + 5*12*500,
    "ai_planning_cost": 4*80,
    "net_benefit": ((12-4)*80 + 5*12*500) - 4*80,
    "roi_percentage": int((((12-4)*80 + 5*12*500) - 4*80) / max(4*80,1) * 100),
    "time_saved_pct": int(((12-4)/12)*100),
    "payback_days": round((4*80) / max(((12-4)*80 + 5*12*500), 1) * 12, 1)
})

st.markdown("<div class='section-label'>💰 ROI & Value Dashboard</div>", unsafe_allow_html=True)
st.markdown("<div class='panel'>", unsafe_allow_html=True)

roi_c1, roi_c2, roi_c3, roi_c4, roi_c5 = st.columns(5)
with roi_c1:
    st.metric("Total Value Delivered", "£{:,}".format(roi_data['total_roi_value']), "Per project cycle")
with roi_c2:
    st.metric("Planning Cost Saved", "£{:,}".format(roi_data['planning_saving']), "{}% faster".format(roi_data['time_saved_pct']))
with roi_c3:
    st.metric("Risk Savings", "£{:,}".format(roi_data['risk_saving']), "{} risks avoided".format(roi_data['risks_identified']))
with roi_c4:
    st.metric("Net Benefit (ROI)", "£{:,}".format(roi_data['net_benefit']), "{}% return".format(roi_data['roi_percentage']))
with roi_c5:
    st.metric("Planning Duration", "{} hrs".format(roi_data['cortex_hours']), "vs {}h manual".format(roi_data['manual_hours']))

planning_bar_pct = min(100, int((roi_data['planning_saving'] / max(roi_data['total_roi_value'], 1)) * 100))
risk_bar_pct = 100 - planning_bar_pct
st.markdown("""
<div style="margin-top:0.9rem;">
  <div style="display:flex;justify-content:space-between;font-size:10px;font-weight:600;color:var(--text-muted);margin-bottom:0.35rem;">
    <span style="font-family:Syne,sans-serif;letter-spacing:0.06em;">ROI Composition</span>
    <span style="color:var(--green-mid);font-weight:700;">£{total_val:,.0f} total value</span>
  </div>
  <div style="height:10px;background:var(--border);border-radius:999px;overflow:hidden;display:flex;">
    <div style="height:100%;width:{plan_pct}%;background:linear-gradient(90deg,#0d4428,#196b41);border-radius:999px 0 0 999px;" title="Planning savings £{plan_sav:,.0f}"></div>
    <div style="height:100%;width:{risk_pct}%;background:linear-gradient(90deg,#c8922a,#e8b84a);" title="Risk savings £{risk_sav:,.0f}"></div>
  </div>
  <div style="display:flex;gap:1.2rem;margin-top:0.4rem;">
    <div style="display:flex;align-items:center;gap:0.35rem;font-size:10px;color:var(--text-secondary);">
      <div style="width:10px;height:10px;background:#196b41;border-radius:2px;flex-shrink:0;"></div>
      Planning savings &nbsp;<strong style="color:var(--green-mid);">£{plan_sav:,.0f}</strong>
    </div>
    <div style="display:flex;align-items:center;gap:0.35rem;font-size:10px;color:var(--text-secondary);">
      <div style="width:10px;height:10px;background:#c8922a;border-radius:2px;flex-shrink:0;"></div>
      Risk avoidance &nbsp;<strong style="color:var(--gold);">£{risk_sav:,.0f}</strong>
    </div>
    <div style="margin-left:auto;font-size:10px;color:var(--text-muted);">
      {manual_hours} manual hrs → {cortex_hours} AI-assisted hr(s) @ £{hrly}/hr · £500 saved/risk/week
    </div>
  </div>
</div>
""".format(
    total_val=roi_data['total_roi_value'],
    plan_pct=planning_bar_pct,
    plan_sav=roi_data['planning_saving'],
    risk_pct=risk_bar_pct,
    risk_sav=roi_data['risk_saving'],
    manual_hours=roi_data['manual_hours'],
    cortex_hours=roi_data['cortex_hours'],
    hrly=roi_data.get('hourly_rate', 80)
), unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)

true_count     = len(past_projects_df)
csv_count      = len(st.session_state.csv_projects)
db_names_set   = set(past_projects_df['project_name'].tolist())
imported_names = st.session_state.imported_csv_names

if true_count > 0 or csv_count > 0:
    col_l, col_r = st.columns([2, 1])
    with col_l:
        st.markdown("<div class='section-label'>Institutional Memory</div>", unsafe_allow_html=True)
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        db_label  = "{} Completed Project{} in Database".format(true_count, 's' if true_count!=1 else '')
        csv_label = " + {} CSV Plan{}".format(csv_count, 's' if csv_count!=1 else '') if csv_count else ""
        st.markdown("<h2>📚 {}{}</h2>".format(db_label, csv_label), unsafe_allow_html=True)
        st.markdown("<div style='font-size:11px;color:var(--text-muted);margin-bottom:0.6rem;'>Cortex selects the most relevant past projects as templates.</div>", unsafe_allow_html=True)

        project_names = past_projects_df['project_name'].tolist()
        csv_display   = []
        csv_map       = {}
        for p in st.session_state.csv_projects:
            nm = p.get("project_name","CSV Project")
            is_in_db = (nm in db_names_set) or (nm in imported_names)
            lbl = "📎 {} ✓ Imported".format(nm) if is_in_db else "📎 {} · Pending Import".format(nm)
            csv_display.append(lbl); csv_map[lbl] = p

        browse_options = ["— Select to view —"] + csv_display + project_names
        if len(browse_options) == 1:
            st.info("No projects loaded yet.")
        else:
            selected_proj = st.selectbox("Browse a project:", browse_options)
            if selected_proj and selected_proj != "— Select to view —":
                if selected_proj in csv_map:
                    proj_data = csv_map[selected_proj]
                    nm = proj_data.get('project_name','CSV Project')
                    is_in_db = (nm in db_names_set) or (nm in imported_names)
                    badge_html = "<span class='csv-badge-imported'>✓ In Database</span>" if is_in_db else "<span class='csv-badge-pending'>Pending Import</span>"
                    with st.expander("📁 {}".format(nm), expanded=True):
                        st.markdown("<div style='font-size:11px;color:var(--text-muted);margin-bottom:.5rem;'>📎 CSV Upload {}</div>".format(badge_html), unsafe_allow_html=True)
                        d1, d2 = st.columns(2)
                        with d1: st.markdown("**Description:** {}".format(proj_data.get('description','N/A')))
                        with d2: st.markdown("**Start:** {} &nbsp; **End:** {}".format(proj_data.get('start_date','—'), proj_data.get('end_date','—')))
                        st.markdown("**WBS Preview:**"); st.code(proj_data.get("wbs_summary","")[:600], language=None)
                else:
                    match = past_projects_df[past_projects_df["project_name"] == selected_proj]
                    if not match.empty:
                        row = match.iloc[0]
                        with st.expander("📁 {}".format(selected_proj), expanded=True):
                            c1, c2 = st.columns(2)
                            with c1:
                                st.markdown("**Description:** {}".format(row.get('description','N/A')))
                                st.markdown("**WBS:** {}...".format(str(row.get('wbs_summary',''))[:300]))
                            with c2:
                                st.markdown("**RAID:** {}...".format(str(row.get('risk_log_summary',''))[:300]))
                                st.markdown("**Deployment:** {}...".format(str(row.get('deployment_plan',''))[:300]))
        st.markdown("</div>", unsafe_allow_html=True)

    with col_r:
        st.markdown("<div class='section-label'>Quick Stats</div>", unsafe_allow_html=True)
        st.markdown("<div class='panel'>", unsafe_allow_html=True)
        st.metric("DB Projects", true_count)
        st.metric("CSV Projects", csv_count)
        st.metric("Total Available", true_count + csv_count)
        st.markdown("</div>", unsafe_allow_html=True)

if generate_button:
    if not sow_pdf or not sow_text.strip():
        st.error("Please upload a SOW PDF before generating the plan.")
        st.stop()

    st.session_state.qa_history     = []
    st.session_state.wbs_structured = None
    st.session_state.plan_methodology = methodology
    st.session_state.plan            = None
    st.session_state.sow_requirements = None
    st.session_state.past_patterns   = None
    st.session_state.risk_lessons    = None
    st.session_state.stakeholder_map = None
    st.session_state.release_notes   = None
    st.session_state.decision_log    = None

    loader = st.empty()

    def show_step(n, msg=""):
        steps = [
            "📄 Analysing SOW — extracting deliverables & scope",
            "🧠 Scanning institutional memory — matching past projects",
            "⚙️  Building WBS from SOW deliverables",
            "📚 Applying past project patterns — RAID, tests, timeline",
            "🔗 Assembling final plan document",
        ]
        items = "".join([
            '<div class="pulse-step' + (' active' if i == n else '') + '">'
            '<span class="step-dot"></span>' + s + '</div>'
            for i, s in enumerate(steps)
        ])
        loader.markdown(
            '<div class="pulse-loader"><div class="pulse-rings"><span></span><span></span><span></span><span></span></div>'
            '<div class="pulse-label">Cortex Project Intel</div>'
            '<div class="pulse-sublabel">' + (msg or "Processing...") + '</div>'
            '<div class="pulse-steps">' + items + '</div></div>',
            unsafe_allow_html=True
        )

    is_agile = any(k in methodology.lower() for k in ['agile','scrum','safe'])
    sprint_len = 2; num_sprints = max(1, duration_weeks // sprint_len)
    compliance_str = ", ".join(compliance_reqs) if compliance_reqs else "None"
    proj_context = (
        "Project: {}\nCustomer: {}\nType: {}\n"
        "Description: {}\nTechnologies: {}\n"
        "Team: {} people | Experience: {}\nRoles: {}\n"
        "Duration: {} weeks | Methodology: {}\n"
        "Delivery Priority: {}\nRisk Appetite: {}\n"
        "Data Sensitivity: {}\nCompliance: {}"
    ).format(new_project_name, customer_name, project_type,
             new_project_desc[:300], new_project_tech,
             team_size, team_experience, team_roster,
             duration_weeks, methodology,
             delivery_priority, risk_appetite,
             has_data_sensitivity, compliance_str)

    show_step(0, "Reading " + sow_pdf.name + " ...")
    sow_call = (
        "You are a senior business analyst reading a Statement of Work (SOW).\n\n"
        "Project context:\n" + proj_context + "\n\nAnalyse the SOW below. Output EXACTLY these 8 numbered sections:\n\n"
        "1. PROJECT OBJECTIVE\n2. KEY DELIVERABLES\n3. ACCEPTANCE CRITERIA\n"
        "4. FUNCTIONAL REQUIREMENTS\n5. NON-FUNCTIONAL REQUIREMENTS\n6. OUT OF SCOPE\n"
        "7. CONSTRAINTS & DEPENDENCIES\n8. RISKS & ASSUMPTIONS IN THE SOW\n\n"
        "SOW TEXT:\n" + sow_text[:5000]
    )
    sow_analysis = cortex_call(sow_call, "SOW Analysis")
    st.session_state.sow_requirements = sow_analysis

    show_step(1, "Scanning past projects...")
    candidates = [{"id": int(r['project_id']), "name": r['project_name'], "desc": str(r.get('description',''))[:80]}
                  for _, r in past_projects_df.iterrows()]
    for ci, cp in enumerate(st.session_state.csv_projects):
        candidates.append({"id": -(ci+1), "name": cp["project_name"], "desc": cp.get("description","")[:80]})

    sel_raw = cortex_call("Select up to 4 past projects most similar to:\n" + proj_context[:400] +
                          "\nCandidates:\n" + json.dumps(candidates[:25])[:1200] +
                          "\nReturn ONLY a JSON array of IDs: [1, 4, 7]", "Template Selection")
    selected_ids = []
    try:
        raw_ids = json.loads(sel_raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip())
        selected_ids = [int(x) for x in raw_ids if str(x).strip().isdigit()][:4]
    except: selected_ids = [c["id"] for c in candidates[:3]]

    pattern_context = ""; used_projects = []
    csv_id_map = {-(i+1): p for i, p in enumerate(st.session_state.csv_projects)}
    for pid in selected_ids:
        match = past_projects_df[past_projects_df['project_id'].astype(str) == str(pid)]
        if pid < 0 and pid in csv_id_map:
            p = csv_id_map[pid]
            pattern_context += "\n-- " + p['project_name'] + " (CSV) --\nWBS: " + p['wbs_summary'][:300] + "\n"
            used_projects.append(p['project_name'])
        elif not match.empty:
            r = match.iloc[0]
            pattern_context += ("\n-- " + r['project_name'] + " --\nRAID: " + str(r.get('risk_log_summary',''))[:280] +
                                "\nTests: " + str(r.get('test_cases_summary',''))[:200] +
                                "\nDeployment: " + str(r.get('deployment_plan',''))[:180] + "\n")
            used_projects.append(r['project_name'])

    if not pattern_context: pattern_context = "No past project data. Use PMO best practices."

    patterns_call = (
        "You are a PMO expert extracting reusable patterns.\n\nProject context:\n" + proj_context +
        "\n\nPast project data:\n" + pattern_context[:3000] +
        "\n\nOutput EXACTLY these four sections:\n"
        "### A. TOP RISKS\nExactly 8 pipe-delimited rows: CATEGORY | TYPE | DESCRIPTION | LIKELIHOOD | IMPACT | MITIGATION | OWNER\n"
        "### B. TEST PATTERNS\nExactly 8 pipe-delimited rows: ID | PRIORITY | TYPE | SCENARIO | PRECONDITIONS | STEPS | EXPECTED RESULT\n"
        "### C. TIMELINE LESSONS\nExactly 5 bullet points with actionable lessons.\n"
        "### D. TEAM STRUCTURE\nRoles as pipe-delimited rows: Role Title | Headcount | Key Responsibilities | Skills Required"
    )
    past_patterns = cortex_call(patterns_call, "Pattern Extraction")
    st.session_state.past_patterns = past_patterns

    show_step(2, "Decomposing SOW into deliverables...")
    sprint_div = ("Divide into " + str(num_sprints) + " Sprints of 2 weeks each." if is_agile else "Divide into phases.")
    wbs_call = (
        "You are a PMO analyst creating a Work Breakdown Structure. "
        "Read the SOW analysis below and produce a detailed sprint plan.\n\n"
        "CRITICAL RULES:\n"
        "- 'stories' must contain SPECIFIC, REAL tasks derived from the SOW deliverables — not generic text.\n"
        "- Each story should be an action: 'Build X', 'Configure Y', 'Test Z', 'Deploy W'.\n"
        "- Use EXACT deliverable names from the SOW analysis in goals and stories.\n"
        "- Produce exactly " + str(num_sprints) + " " + ("sprints" if is_agile else "phases") + ".\n"
        "- Do NOT use placeholder text like 'Task 1', 'Define', 'Build', 'Test' as standalone items.\n\n"
        "Project: " + new_project_name + " | Customer: " + customer_name + "\n"
        "Technologies: " + new_project_tech + "\n"
        "Team: " + team_roster + "\n\n"
        "SOW Analysis:\n" + sow_analysis[:2800] + "\n\n"
        + sprint_div + "\n\n"
        "Return ONLY a valid JSON array (no markdown, no extra text):\n"
        '[{"sprint_num":1,"name":"Sprint 1: <specific title>","weeks":"Week 1-2",'
        '"goal":"<specific deliverable from SOW>","sow_ref":"<SOW section referenced>",'
        '"stories":["<specific task 1 from SOW>","<specific task 2>","<specific task 3>","<specific task 4>"],'
        '"deliverable":"<specific acceptance criterion from SOW>"}]'
    )
    raw_wbs = cortex_call(wbs_call, "WBS Generation")
    wbs_sprints_gen = []
    if raw_wbs:
        try:
            clean = raw_wbs.strip()
            for fence in ["```json", "```JSON", "```"]:
                if clean.startswith(fence): clean = clean[len(fence):]
                if clean.endswith(fence):   clean = clean[:-len(fence)]
            clean = clean.strip()
            start = clean.find('['); end = clean.rfind(']')
            if start != -1 and end != -1: clean = clean[start:end+1]
            wbs_sprints_gen = json.loads(clean)
            if not isinstance(wbs_sprints_gen, list):
                wbs_sprints_gen = []
            else:
                placeholders = {'define','build','test','task 1','task 2','task 3','task 4',
                                'specific task 1','specific task 2','tbd','n/a',''}
                total_real = sum(
                    1 for sp in wbs_sprints_gen
                    for story in sp.get('stories', [])
                    if str(story).lower().strip() not in placeholders and len(str(story)) > 10
                )
                if total_real < len(wbs_sprints_gen) * 2:
                    wbs_sprints_gen = []
        except Exception:
            wbs_sprints_gen = []

    if not wbs_sprints_gen:
        wbs_sprints_gen = parse_wbs_into_sprints(sow_analysis, duration_weeks, methodology)
    st.session_state.wbs_structured = wbs_sprints_gen

    show_step(3, "Applying past project lessons...")
    wbs_for_prompt = "\n".join([
        "  " + s.get('name','') + " (" + s.get('weeks','') + "): " + s.get('goal','') + " => " + s.get('deliverable','')
        for s in wbs_sprints_gen
    ])[:1200]

    show_step(4, "Assembling final plan document...")
    assembly_call = (
        "You are a senior PMO consultant writing the final project plan.\n\n"
        "Project context:\n" + proj_context + "\n\n"
        "=== SOW SCOPE ===\n" + sow_analysis[:1800] + "\n\n"
        "=== WBS ===\n" + wbs_for_prompt + "\n\n"
        "=== PAST PROJECT PATTERNS ===\n" + past_patterns[:1200] + "\n\n"
        "Write EXACTLY these 7 sections with EXACT ## headers:\n\n"
        "## EXECUTIVE SUMMARY\n3 paragraphs covering deliverables, past patterns, and KPIs.\n\n"
        "## RAID LOG\nMin 10 pipe-delimited rows: CATEGORY | TYPE | DESCRIPTION | LIKELIHOOD | IMPACT | MITIGATION | OWNER\n\n"
        "## TEST CASES\nMin 10 pipe-delimited rows: ID | PRIORITY | TYPE | SCENARIO | PRECONDITIONS | STEPS | EXPECTED RESULT\n\n"
        "## DELIVERY TIMELINE\nPipe-delimited rows: WEEK | SPRINT/PHASE | MILESTONE | OWNER | STATUS | RISK FLAG\n\n"
        "## DEPLOYMENT STRATEGY\nProse. Phased go-live, rollback, sign-off gates.\n\n"
        "## SOW TRACEABILITY MATRIX\nPipe-delimited rows: SOW REQUIREMENT | SPRINT/PHASE | KEY TASKS | ACCEPTANCE CRITERIA | STATUS\n\n"
        "## LESSONS FROM PAST PROJECTS\nBullets. Each with past project reference and change made to this plan.\n\n"
        "Sections with pipe-delimited format: data rows ONLY, no headers, no blanks within."
    )
    full_plan = cortex_call(assembly_call, "Plan Assembly")
    st.session_state.plan      = full_plan or "Plan assembly failed. SOW analysis and WBS still available."
    st.session_state.plan_name = new_project_name
    loader.empty(); st.rerun()

if st.session_state.plan:
    plan      = st.session_state.plan
    plan_name = st.session_state.plan_name
    meth      = st.session_state.plan_methodology or "Agile / Scrum"
    sprints   = st.session_state.wbs_structured or []
    is_agile  = any(k in meth.lower() for k in ['agile','scrum','safe'])

    st.markdown("<div class='section-label'>Delivery Impact</div>", unsafe_allow_html=True)
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    mc1,mc2,mc3,mc4,mc5 = st.columns(5)
    with mc1: st.metric("Planning Time", "Weeks → Minutes", "90% reduction")
    with mc2: st.metric("Risk Identification", "Reactive → Predictive", "From past projects")
    with mc3: st.metric("Templates Used", min(len(past_projects_df), 4), "matched by Cortex")
    with mc4: st.metric("Sprint / Phase Count", len(sprints) if sprints else "—")
    with mc5: st.metric("Plan Sections", "7 structured", "RAID · Tests · Timeline")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<div class='section-label' style='margin-top:0.25rem;'>Generated Project Plan</div>", unsafe_allow_html=True)

    tab_plan,tab_wbs,tab_sow,tab_qa,tab_risks,tab_stakeholder,tab_roi,tab_export = st.tabs([
        "📋 Full Plan","🏃 WBS / Sprints","📄 SOW Analysis","💬 Ask the Plan",
        "⚠️ Risk Analyser","👥 Stakeholder Map","💰 ROI Report","📤 Export"])

    with tab_plan:
        st.markdown("<h2>📋 {}</h2>".format(plan_name), unsafe_allow_html=True)

        def render_plan_with_tables(plan_text):
            LH_BG  = {"High":"rgba(192,57,43,0.09)","Medium":"rgba(200,146,42,0.09)","Low":"rgba(25,107,65,0.09)"}
            LH_COL = {"High":"#c0392b","Medium":"#c8922a","Low":"#196b41"}
            CAT_BG  = {"Risk":"rgba(192,57,43,0.08)","Assumption":"rgba(26,94,168,0.08)",
                       "Issue":"rgba(200,146,42,0.08)","Dependency":"rgba(107,47,168,0.08)"}
            CAT_COL = {"Risk":"#c0392b","Assumption":"#1a5ea8","Issue":"#c8922a","Dependency":"#6b2fa8"}
            TT_BG  = {"Integration":"rgba(26,94,168,0.08)","Data Quality":"rgba(25,107,65,0.08)",
                      "Security":"rgba(192,57,43,0.08)","Performance":"rgba(200,146,42,0.08)",
                      "Functional":"rgba(107,47,168,0.08)","Regression":"rgba(100,116,139,0.08)",
                      "Uat":"rgba(25,107,65,0.08)","UAT":"rgba(25,107,65,0.08)"}
            TT_COL = {"Integration":"#1a5ea8","Data Quality":"#196b41","Security":"#c0392b",
                      "Performance":"#c8922a","Functional":"#6b2fa8","Regression":"#64748b","Uat":"#196b41","UAT":"#196b41"}
            ST_BG  = {"To Do":"rgba(100,116,139,0.08)","Planned":"rgba(200,146,42,0.08)",
                      "In Progress":"rgba(26,94,168,0.08)","Done":"rgba(25,107,65,0.08)"}
            ST_COL = {"To Do":"#64748b","Planned":"#c8922a","In Progress":"#1a5ea8","Done":"#196b41"}

            SECTIONS = {
                "EXECUTIVE SUMMARY":  {"icon":"📋","col":"#196b41","table":False,"headers":[]},
                "RAID LOG":           {"icon":"⚠️","col":"#c0392b","table":True,
                    "headers":["Category","Type","Description","Likelihood","Impact","Mitigation","Owner"],
                    "colfn": lambda ci, v: (
                        "background:%s;color:%s;font-weight:700;text-align:center;" % (CAT_BG.get(v.title(),"rgba(0,0,0,0.03)"), CAT_COL.get(v.title(),"var(--text-secondary)")) if ci==0 else
                        "background:rgba(0,0,0,0.02);color:var(--text-muted);text-align:center;font-size:10px;" if ci==1 else
                        "background:rgba(0,0,0,0.02);color:var(--text-primary);font-weight:600;" if ci==2 else
                        "background:%s;color:%s;font-weight:700;text-align:center;" % (LH_BG.get(v.title(),"rgba(0,0,0,0.03)"), LH_COL.get(v.title(),"var(--text-muted)")) if ci in (3,4) else
                        "background:rgba(0,0,0,0.02);color:var(--text-secondary);" if ci==5 else
                        "background:var(--green-xlight);color:var(--green-mid);font-weight:600;text-align:center;"
                    )},
                "TEST CASES":         {"icon":"🧪","col":"#1a5ea8","table":True,
                    "headers":["ID","Priority","Type","Scenario","Preconditions","Steps","Expected Result"],
                    "colfn": lambda ci, v: (
                        "background:var(--blue-bg);color:var(--blue);font-weight:700;text-align:center;white-space:nowrap;" if ci==0 else
                        "background:%s;color:%s;font-weight:700;text-align:center;" % (LH_BG.get(v.title(),"rgba(0,0,0,0.03)"), LH_COL.get(v.title(),"var(--text-muted)")) if ci==1 else
                        "background:%s;color:%s;font-weight:700;text-align:center;white-space:nowrap;" % (TT_BG.get(v.title(),TT_BG.get(v,"rgba(0,0,0,0.03)")), TT_COL.get(v.title(),TT_COL.get(v,"var(--text-muted)"))) if ci==2 else
                        "background:rgba(0,0,0,0.02);color:var(--text-primary);font-weight:600;" if ci==3 else
                        "background:rgba(0,0,0,0.02);color:var(--text-secondary);" if ci in (4,5) else
                        "background:var(--green-xlight);color:var(--green-mid);font-weight:600;"
                    )},
                "DELIVERY TIMELINE":  {"icon":"📅","col":"#c8922a","table":True,
                    "headers":["Week","Sprint / Phase","Milestone","Owner","Status","Risk Flag"],
                    "colfn": lambda ci, v: (
                        "background:var(--gold-bg);color:var(--gold);font-weight:700;text-align:center;white-space:nowrap;" if ci==0 else
                        "background:var(--blue-bg);color:var(--blue);font-weight:600;" if ci==1 else
                        "background:rgba(0,0,0,0.02);color:var(--text-primary);font-weight:600;" if ci==2 else
                        "background:var(--green-xlight);color:var(--green-mid);text-align:center;" if ci==3 else
                        "background:%s;color:%s;font-weight:700;text-align:center;" % (ST_BG.get(v.title(),ST_BG.get(v,"rgba(0,0,0,0.03)")), ST_COL.get(v.title(),ST_COL.get(v,"var(--text-muted)"))) if ci==4 else
                        "background:%s;color:%s;font-weight:700;text-align:center;" % (LH_BG.get(v.title(),LH_BG.get(v,"var(--green-xlight)")), LH_COL.get(v.title(),LH_COL.get(v,"var(--green-mid)")))
                    )},
                "DEPLOYMENT STRATEGY":{"icon":"🚀","col":"#6b2fa8","table":False,"headers":[]},
                "SOW TRACEABILITY":   {"icon":"🔗","col":"#196b41","table":True,
                    "headers":["SOW Requirement","Sprint / Phase","Key Tasks","Acceptance Criteria","Status"],
                    "colfn": lambda ci, v: (
                        "background:var(--green-xlight);color:var(--green-mid);font-weight:700;" if ci==0 else
                        "background:var(--blue-bg);color:var(--blue);font-weight:600;text-align:center;white-space:nowrap;" if ci==1 else
                        "background:rgba(0,0,0,0.02);color:var(--text-secondary);" if ci==2 else
                        "background:rgba(0,0,0,0.02);color:var(--text-muted);font-style:italic;" if ci==3 else
                        "background:%s;color:%s;font-weight:700;text-align:center;" % (ST_BG.get(v.title(),ST_BG.get(v,"rgba(0,0,0,0.03)")), ST_COL.get(v.title(),ST_COL.get(v,"var(--text-muted)")))
                    )},
                "LESSONS FROM PAST":  {"icon":"📚","col":"#196b41","table":False,"headers":[]},
            }

            def make_table(sec_key, rows):
                if not rows:
                    return ""
                sec = SECTIONS.get(sec_key, {})
                headers = sec.get("headers", [])
                colfn = sec.get("colfn", lambda ci, v: "background:rgba(0,0,0,0.02);color:var(--text-secondary);")
                accent = sec.get("col","#196b41")
                th = "".join(
                    f"<th style='padding:0.55rem 0.9rem;text-align:left;font-size:9px;font-weight:800;"
                    f"letter-spacing:0.09em;color:{accent};white-space:nowrap;background:var(--green-xlight);"
                    f"border-bottom:2px solid {accent}44;border-right:1px solid var(--border);"
                    f"border-top:1px solid var(--border);border-left:1px solid var(--border);'>{h}</th>"
                    for h in headers
                )
                body = ""
                for ri, row in enumerate(rows):
                    while len(row) < len(headers):
                        row.append("")
                    row = row[:len(headers)]
                    cells = ""
                    row_bg = "rgba(0,0,0,0)" if ri % 2 == 0 else "rgba(0,0,0,0.015)"
                    for ci, cell in enumerate(row):
                        v = str(cell).strip()
                        sty = colfn(ci, v)
                        cells += f"<td style='padding:0.48rem 0.9rem;font-size:11px;vertical-align:top;line-height:1.65;border-bottom:1px solid var(--border);border-right:1px solid var(--border);background:{row_bg};{sty}'>{v}</td>"
                    body += f"<tr>{cells}</tr>"
                return (
                    "<div style='overflow-x:auto;margin:0.7rem 0 1.6rem 0;border-radius:10px;"
                    "border:1.5px solid %s33;box-shadow:var(--shadow-sm);'>"
                    "<table style='width:100%%;border-collapse:collapse;'>"
                    "<thead>%s</thead><tbody>%s</tbody></table></div>"
                ) % (accent, th, body)

            def make_header(sec_key, title):
                sec = SECTIONS.get(sec_key, {"icon":"▸","col":"#196b41"})
                return (
                    "<div style='display:flex;align-items:center;gap:0.6rem;"
                    "margin:1.6rem 0 0.7rem 0;padding:0.6rem 1rem;"
                    "background:linear-gradient(135deg,rgba(%s),rgba(%s,0.03));"
                    "border-radius:10px;border-left:3px solid %s;'>"
                    "<span style='font-size:1.1rem;'>%s</span>"
                    "<span style='font-size:12px;font-weight:800;color:%s;"
                    "text-transform:uppercase;letter-spacing:0.12em;font-family:Syne,sans-serif;'>%s</span>"
                    "</div>"
                ) % (
                    _hex_to_rgb(sec["col"]) + ",0.06",
                    _hex_to_rgb(sec["col"]),
                    sec["col"], sec["icon"], sec["col"], title
                )

            def _hex_to_rgb(h):
                h = h.lstrip('#')
                try:
                    return ','.join(str(int(h[i:i+2],16)) for i in (0,2,4))
                except:
                    return "0,0,0"

            lines = plan_text.split("\n")
            output = ""
            cur_sec = None
            tbl_rows = []
            in_tbl = False
            for line in lines:
                s = line.strip()
                if s.startswith("##") or s.startswith("# "):
                    if in_tbl and tbl_rows and cur_sec:
                        output += make_table(cur_sec, tbl_rows)
                    tbl_rows = []
                    in_tbl = False
                    cur_sec = None
                    title = s.lstrip("#").strip()
                    matched = next((k for k in SECTIONS if k in title.upper()), None)
                    cur_sec = matched
                    output += make_header(matched or title.upper(), title)
                    if matched and SECTIONS[matched]["table"]:
                        in_tbl = True
                elif in_tbl and "|" in s:
                    parts = [p.strip() for p in s.split("|") if p.strip()]
                    if parts and not all(set(p) <= set("-: ") for p in parts):
                        hdrs = SECTIONS.get(cur_sec, {}).get("headers", [])
                        if parts[0].upper() not in [h.upper() for h in hdrs]:
                            tbl_rows.append(parts)
                elif in_tbl and s == "":
                    if tbl_rows and cur_sec:
                        output += make_table(cur_sec, tbl_rows)
                        tbl_rows = []
                    in_tbl = False
                elif not in_tbl and s:
                    if s.startswith("- ") or s.startswith("* ") or s.startswith("• "):
                        c = re.sub(r"^[-*•]\s*", "", s)
                        output += (
                            "<div style='display:flex;gap:0.55rem;padding:0.25rem 0.2rem;"
                            "color:var(--text-primary);font-size:12px;line-height:1.75;'>"
                            "<span style='color:var(--green-bright);flex-shrink:0;font-size:0.7rem;margin-top:0.2rem;'>◆</span>"
                            "<span>%s</span></div>" % c
                        )
                    elif re.match(r"^\d+\.\s", s):
                        m2 = re.match(r"^(\d+)\.\s*(.*)", s)
                        if m2:
                            output += (
                                "<div style='display:flex;gap:0.55rem;padding:0.25rem 0.2rem;"
                                "color:var(--text-primary);font-size:12px;line-height:1.75;'>"
                                "<span style='color:var(--green-mid);font-weight:800;flex-shrink:0;"
                                "min-width:1.4rem;font-family:Syne,sans-serif;'>%s.</span><span>%s</span></div>"
                                % (m2.group(1), m2.group(2))
                            )
                    elif s.startswith("**") or s.startswith("__"):
                        clean = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', s)
                        output += "<div style='color:var(--text-primary);font-size:12px;line-height:1.85;padding:0.12rem 0.2rem;font-weight:600;'>%s</div>" % clean
                    else:
                        output += "<div style='color:var(--text-secondary);font-size:12px;line-height:1.9;padding:0.12rem 0.2rem;'>%s</div>" % s
            if in_tbl and tbl_rows and cur_sec:
                output += make_table(cur_sec, tbl_rows)
            return output

        rendered = render_plan_with_tables(plan)
        st.markdown("<div class='plan-content-wrap'>" + rendered + "</div>", unsafe_allow_html=True)
        st.success("✅ Plan generated — WBS from SOW · RAID & Tests from institutional memory.")

    with tab_wbs:
        label = "Sprint Board" if is_agile else "Work Breakdown Structure"
        st.markdown("<h2>🏃 {} — {}</h2>".format(label, plan_name), unsafe_allow_html=True)

        placeholders = {'define','build','test','task 1','task 2','task 3','tbd','n/a','iteration 1'}
        def is_thin(sprint_list):
            if not sprint_list: return True
            total = sum(len(sp.get('stories',[])) for sp in sprint_list)
            real  = sum(
                1 for sp in sprint_list for s in sp.get('stories',[])
                if str(s).lower().strip() not in placeholders and len(str(s)) > 10
            )
            return total == 0 or real < total * 0.5

        if is_thin(sprints):
            with st.spinner("Extracting detailed tasks from plan…"):
                sprints = parse_wbs_into_sprints(plan, duration_weeks if 'duration_weeks' in dir() else 12, meth)
                st.session_state.wbs_structured = sprints

        if not sprints:
            st.info("No sprint data — regenerate the plan.")
        else:
            sum_cols = st.columns(min(len(sprints), 6))
            sp_colors = ["#0d4428","#196b41","#228b55","#1a5ea8","#6b2fa8","#c8922a"]
            for i, sp in enumerate(sprints[:6]):
                with sum_cols[i]:
                    sp_name_short = sp.get('name','').split(':')[0] if ':' in sp.get('name','') else sp.get('name','')
                    col_h = sp_colors[i % len(sp_colors)]
                    st.markdown(
                        "<div style='background:{col};border-radius:10px;padding:0.65rem 0.7rem;"
                        "text-align:center;box-shadow:var(--shadow-sm);'>"
                        "<div style='font-size:9px;font-weight:700;color:rgba(255,255,255,0.6);"
                        "text-transform:uppercase;letter-spacing:0.12em;font-family:Syne,sans-serif;'>{wks}</div>"
                        "<div style='font-size:11px;font-weight:700;color:#ffffff;margin-top:0.2rem;line-height:1.3;'>{nm}</div>"
                        "</div>".format(col=col_h, wks=sp.get('weeks',''), nm=sp_name_short),
                        unsafe_allow_html=True
                    )
            st.markdown("<hr>", unsafe_allow_html=True)

            scol1, scol2 = st.columns(2)
            task_status_colors = ["#196b41","#1a5ea8","#6b2fa8","#c8922a","#c0392b","#0e7c6e"]
            for i, sp in enumerate(sprints):
                raw_tasks = sp.get("stories", sp.get("tasks", []))
                display_tasks = []
                for t in raw_tasks:
                    val = str(t.get("name", t.get("text", str(t)))) if isinstance(t, dict) else str(t)
                    val = val.strip()
                    if val and val.lower() not in placeholders and len(val) > 3:
                        display_tasks.append(val)
                with (scol1 if i % 2 == 0 else scol2):
                    deliverable = str(sp.get("deliverable", ""))
                    sow_ref     = str(sp.get("sow_ref", ""))
                    sp_name     = str(sp.get("name", ""))
                    sp_weeks    = str(sp.get("weeks", ""))
                    sp_goal     = str(sp.get("goal", ""))
                    border_col  = sp_colors[i % len(sp_colors)]

                    card = (
                        "<div style='background:#ffffff;border:1.5px solid var(--border);"
                        "border-radius:var(--radius-md);padding:0;margin-bottom:0.7rem;"
                        "border-left:3px solid {col};box-shadow:var(--shadow-sm);overflow:hidden;'>"
                        "<div style='background:{col};padding:0.75rem 1rem;'>"
                        "<div style='font-size:11px;font-weight:800;color:#ffffff;font-family:Syne,sans-serif;'>{nm}</div>"
                        "<div style='font-size:10px;color:rgba(255,255,255,0.7);margin-top:0.15rem;'>{wks}</div>"
                        "</div>"
                        "<div style='padding:0.85rem 1rem;'>"
                    ).format(col=border_col, nm=sp_name, wks=sp_weeks)

                    if sow_ref:
                        card += "<div style='font-size:9px;color:var(--text-muted);font-weight:600;margin-bottom:0.35rem;text-transform:uppercase;letter-spacing:0.08em;font-family:Syne,sans-serif;'>📎 " + sow_ref + "</div>"
                    card += "<div style='font-size:11px;color:var(--text-primary);font-weight:600;margin-bottom:0.6rem;padding:0.4rem 0.6rem;background:var(--green-xlight);border-radius:6px;'>🎯 " + sp_goal + "</div>"

                    if display_tasks:
                        for ti, t in enumerate(display_tasks):
                            dot_col = task_status_colors[ti % len(task_status_colors)]
                            card += (
                                "<div style='font-size:11px;color:var(--text-secondary);padding:0.25rem 0;"
                                "border-bottom:1px solid var(--border);display:flex;gap:0.5rem;align-items:flex-start;'>"
                                "<span style='width:6px;height:6px;border-radius:50%;background:{col};"
                                "flex-shrink:0;margin-top:0.25rem;'></span>"
                                "<span>{task}</span></div>"
                            ).format(col=dot_col, task=t)
                    else:
                        card += "<div style='font-size:11px;color:var(--text-disabled);font-style:italic;'>No tasks extracted</div>"

                    if deliverable:
                        card += (
                            "<div style='font-size:10px;color:{col};margin-top:0.55rem;font-weight:700;"
                            "padding:0.35rem 0.6rem;background:{col}18;border-radius:6px;"
                            "border:1px solid {col}33;'>✓ {deliv}</div>"
                        ).format(col=border_col, deliv=deliverable)

                    card += "</div></div>"
                    st.markdown(card, unsafe_allow_html=True)

    with tab_sow:
        st.markdown("<h2>📄 SOW Analysis</h2>", unsafe_allow_html=True)
        sow_req = st.session_state.sow_requirements or ""
        if sow_req:
            SOW_ICONS = {
                "PROJECT OBJECTIVE": ("🎯","#196b41"), "KEY DELIVERABLES": ("📦","#196b41"),
                "ACCEPTANCE CRITERIA": ("✅","#196b41"), "FUNCTIONAL REQUIREMENTS": ("⚙️","#1a5ea8"),
                "NON-FUNCTIONAL REQUIREMENTS": ("🔧","#1a5ea8"), "OUT OF SCOPE": ("🚫","#c8922a"),
                "CONSTRAINTS": ("⛓️","#c8922a"), "RISKS & ASSUMPTIONS": ("⚠️","#c0392b"),
                "RISKS AND ASSUMPTIONS": ("⚠️","#c0392b"),
            }
            def render_sow_analysis(text):
                lines = text.split('\n'); html = ""; in_section = False
                for line in lines:
                    s = line.strip()
                    if not s:
                        if in_section: html += "<div style='height:0.3rem;'></div>"
                        continue
                    is_header = False; header_text = ""
                    if re.match(r'^\d+\.', s): header_text = re.sub(r'^\d+\.\s*','',s).strip(); is_header = True
                    elif s.startswith('##') or s.startswith('# '): header_text = s.lstrip('#').strip(); is_header = True
                    if is_header:
                        if in_section: html += "</div></div>"
                        ht_upper = header_text.upper()
                        icon, color = "📌", "#196b41"
                        for key, (ic, co) in SOW_ICONS.items():
                            if key in ht_upper: icon, color = ic, co; break
                        r,g,b = (int(color.lstrip('#')[i:i+2],16) for i in (0,2,4)) if color.startswith('#') else (25,107,65)
                        html += (
                            f"<div style='margin-bottom:1rem;'>"
                            f"<div style='display:flex;align-items:center;gap:0.6rem;padding:0.65rem 1rem;"
                            f"margin-bottom:0.55rem;background:rgba({r},{g},{b},0.07);"
                            f"border-radius:10px;border-left:3px solid {color};'>"
                            f"<span style='font-size:1.05rem;'>{icon}</span>"
                            f"<span style='font-size:12px;font-weight:800;color:{color};"
                            f"text-transform:uppercase;letter-spacing:0.10em;font-family:Syne,sans-serif;'>{header_text}</span></div>"
                            f"<div style='padding-left:0.5rem;'>"
                        )
                        in_section = True
                    elif s.startswith('-') or s.startswith('•') or s.startswith('*'):
                        content = re.sub(r'^[-•*]\s*','',s)
                        html += (
                            f"<div style='display:flex;gap:0.55rem;padding:0.25rem 0;"
                            f"color:var(--text-primary);font-size:12px;line-height:1.7;'>"
                            f"<span style='color:var(--green-bright);flex-shrink:0;font-size:0.65rem;margin-top:0.25rem;'>◆</span>"
                            f"<span>{content}</span></div>"
                        )
                    else:
                        html += f"<div style='color:var(--text-secondary);font-size:12px;line-height:1.85;padding:0.12rem 0;'>{s}</div>"
                if in_section: html += "</div></div>"
                return html
            st.markdown("<div class='plan-content-wrap'>" + render_sow_analysis(sow_req) + "</div>", unsafe_allow_html=True)
        else:
            st.info("SOW analysis not available — regenerate the plan.")

    with tab_qa:
        st.markdown("<h2>💬 Ask Cortex About This Plan</h2>", unsafe_allow_html=True)
        suggestions = [
            "What are the top 3 risks?","Write a kick-off email.",
            "Sprint 1 in detail?","Generate a RACI matrix.",
            "What compliance tasks are needed?","Rewrite exec summary simply.",
            "What is the critical path?","Suggest risk mitigations."
        ]
        sc1,sc2,sc3 = st.columns(3)
        for i, s in enumerate(suggestions):
            with [sc1,sc2,sc3][i%3]:
                if st.button(s, key="sugg_{}".format(i)): st.session_state['pending_question'] = s
        st.markdown("<hr>", unsafe_allow_html=True)
        if 'qa_input_value' not in st.session_state: st.session_state['qa_input_value'] = ''
        if st.session_state.get('pending_question'): st.session_state['qa_input_value'] = st.session_state.pop('pending_question')
        user_q = st.text_input("Your question:", value=st.session_state['qa_input_value'], placeholder="e.g. What dependencies could block Sprint 2?")
        if st.button("🤖 Ask Cortex", key="ask_btn"):
            q = (user_q or '').strip()
            if not q: st.warning("Please type a question first.")
            else:
                sow_ctx = (st.session_state.sow_requirements or "")[:600]
                wbs_ctx = "\n".join([s.get('name','') + " (" + s.get('weeks','') + "): " + s.get('goal','') for s in (st.session_state.wbs_structured or [])])[:500]
                qa_prompt = ("You are a senior PMO consultant. Answer using the project context.\n\nProject: " + plan_name +
                             "\nSOW:\n" + sow_ctx + "\n\nWBS:\n" + wbs_ctx + "\n\nPlan:\n" + plan[:1800] + "\n\nQuestion: " + q +
                             "\n\nAnswer specifically and practically.")
                with st.spinner("Cortex is thinking…"):
                    ans = cortex_call(qa_prompt, "Q&A")
                if ans:
                    st.session_state.qa_history.append({"q": q, "a": ans})
                    st.session_state['qa_input_value'] = ''; st.rerun()
                else: st.error("No response.")
        if st.session_state.qa_history:
            st.markdown(
                "<div style='font-size:9px;font-weight:700;color:var(--text-muted);text-transform:uppercase;"
                "letter-spacing:0.14em;margin:0.75rem 0 0.5rem;border-bottom:1px solid var(--border);"
                "padding-bottom:0.3rem;font-family:Syne,sans-serif;'>💬 Conversation History</div>",
                unsafe_allow_html=True
            )
            for qa in reversed(st.session_state.qa_history):
                st.markdown(
                    "<div style='background:var(--green-xlight);border:1.5px solid var(--border-strong);"
                    "border-left:3px solid var(--green-mid);border-radius:10px;padding:0.7rem 1rem;"
                    "margin-bottom:0.3rem;'><div style='font-size:9px;font-weight:700;color:var(--text-muted);"
                    "text-transform:uppercase;margin-bottom:0.2rem;font-family:Syne,sans-serif;'>You asked</div>"
                    "<div style='font-size:12px;color:var(--text-primary);font-weight:600;'>❓ " + qa['q'] + "</div></div>",
                    unsafe_allow_html=True
                )
                st.markdown(
                    "<div style='background:var(--bg-panel-alt);border:1.5px solid var(--border);"
                    "border-left:3px solid var(--blue);border-radius:0 10px 10px 10px;padding:0.9rem 1.1rem;"
                    "margin-bottom:1rem;'><div style='font-size:9px;font-weight:700;color:var(--blue);"
                    "text-transform:uppercase;margin-bottom:0.5rem;font-family:Syne,sans-serif;'>🤖 Cortex</div>"
                    "<div style='font-size:12px;color:var(--text-primary);line-height:1.85;'>"
                    + qa['a'].replace('\n','<br>') + "</div></div>",
                    unsafe_allow_html=True
                )
            if st.button("🗑️ Clear History", key="clear_qa"): st.session_state.qa_history = []; st.rerun()

    with tab_risks:
        st.markdown("<h2>⚠️ Risk Analyser</h2>", unsafe_allow_html=True)

        if st.session_state.current_risks:
            risks_for_heatmap = st.session_state.current_risks
            hm_data = {}
            for r in risks_for_heatmap:
                lh = r.get('likelihood','Medium').strip().title()
                im = r.get('impact','Medium').strip().title()
                hm_data[(lh,im)] = hm_data.get((lh,im), 0) + 1

            levels = ["Low","Medium","High"]
            hm_cells = ""
            for lh in reversed(levels):
                hm_cells += f"<div style='display:flex;gap:0.3rem;margin-bottom:0.3rem;align-items:center;'><div style='font-size:9px;color:var(--text-muted);width:48px;text-align:right;padding-right:0.4rem;font-weight:600;'>{lh}</div>"
                for im in levels:
                    cnt = hm_data.get((lh,im), 0)
                    if lh == "High" and im == "High":
                        bg,tc = "#c0392b","#fff"
                    elif (lh == "High" and im == "Medium") or (lh == "Medium" and im == "High"):
                        bg,tc = "#c8922a","#fff"
                    elif lh == "Low" and im == "Low":
                        bg,tc = "#196b41","#fff"
                    else:
                        bg,tc = "#e4f5ec","#196b41"
                    hm_cells += f"<div style='width:52px;height:44px;background:{bg};border-radius:8px;display:flex;align-items:center;justify-content:center;flex-direction:column;'><span style='font-size:16px;font-weight:900;color:{tc};line-height:1;'>{cnt if cnt else '·'}</span></div>"
                hm_cells += "</div>"

            st.markdown(
                "<div style='background:#ffffff;border:1.5px solid var(--border);border-radius:14px;"
                "padding:1.1rem 1.3rem;margin-bottom:1.1rem;box-shadow:var(--shadow-sm);'>"
                "<div style='font-size:10px;font-weight:800;color:var(--text-muted);text-transform:uppercase;"
                "letter-spacing:0.14em;margin-bottom:0.9rem;font-family:Syne,sans-serif;'>📊 Risk Heatmap</div>"
                "<div style='display:flex;gap:0.5rem;'>"
                "<div style='display:flex;flex-direction:column;gap:0;'>" + hm_cells + "</div>"
                "<div style='margin-left:0.5rem;'>"
                "<div style='display:flex;gap:0.3rem;margin-bottom:0.25rem;'>"
                + "".join(f"<div style='width:52px;font-size:9px;font-weight:700;color:var(--text-muted);text-align:center;'>{im}</div>" for im in levels) +
                "</div>"
                "<div style='font-size:9px;color:var(--text-muted);margin-top:0.4rem;'>← Impact →</div>"
                "</div></div>"
                "<div style='display:flex;gap:0.6rem;margin-top:0.7rem;flex-wrap:wrap;'>"
                "<span style='display:flex;align-items:center;gap:0.3rem;font-size:10px;color:#c0392b;font-weight:700;'><div style='width:10px;height:10px;background:#c0392b;border-radius:2px;'></div>Critical</span>"
                "<span style='display:flex;align-items:center;gap:0.3rem;font-size:10px;color:#c8922a;font-weight:700;'><div style='width:10px;height:10px;background:#c8922a;border-radius:2px;'></div>High</span>"
                "<span style='display:flex;align-items:center;gap:0.3rem;font-size:10px;color:#196b41;font-weight:700;'><div style='width:10px;height:10px;background:#196b41;border-radius:2px;'></div>Low</span>"
                "</div></div>",
                unsafe_allow_html=True
            )

        if st.button("🔍 Extract & Score Risks with Cortex"):
            rp = ("Extract all risks from this plan for \"{}\".\nPlan:\n{}\n\n"
                  "Return ONLY valid JSON array. For each risk, provide:\n"
                  "- 'risk': title\n- 'description': detailed description (max 100 chars)\n"
                  "- 'type': Technical/Resource/Timeline/Budget/Compliance/Process\n"
                  "- 'likelihood': High/Medium/Low\n- 'impact': High/Medium/Low\n"
                  "- 'mitigation': one sentence\n- 'owner': role (e.g., PM, Platform Engineer)\n"
                  "Example: [{{\"risk\":\"Data access permissions\",\"description\":\"Missing IAM roles\","
                  "\"type\":\"Technical\",\"likelihood\":\"High\",\"impact\":\"High\","
                  "\"mitigation\":\"Pre-provision roles before Sprint 1\",\"owner\":\"Platform Engineer\"}}]"
                  ).format(plan_name, plan[:2500])
            with st.spinner("Analyzing risks..."):
                raw = cortex_call(rp, "Risk Analyser")
                try:
                    raw = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
                    risks = json.loads(raw)
                    if not isinstance(risks, list): risks = []
                except:
                    risks = [
                        {"risk":"Data access permissions","description":"Missing IAM roles for Snowflake","type":"Technical","likelihood":"High","impact":"High","mitigation":"Pre-provision IAM roles before Sprint 1","owner":"Platform Engineer"},
                        {"risk":"Scope creep","description":"Unexpected stakeholder change requests","type":"Timeline","likelihood":"Medium","impact":"High","mitigation":"Lock scope after Sprint 2","owner":"PM"},
                        {"risk":"Key person dependency","description":"Single architect owns critical design","type":"Resource","likelihood":"Medium","impact":"Medium","mitigation":"Cross-train 2 members","owner":"Tech Lead"}
                    ]
            st.session_state.current_risks = risks

            lbg={"High":"rgba(192,57,43,0.08)","Medium":"rgba(200,146,42,0.08)","Low":"rgba(25,107,65,0.08)"}
            lco={"High":"#c0392b","Medium":"#c8922a","Low":"#196b41"}
            type_icons = {"Technical":"⚙️","Resource":"👥","Timeline":"📅","Budget":"💰","Compliance":"🔒","Process":"📋"}
            for r in risks:
                lh=r.get('likelihood','Medium'); im=r.get('impact','Medium')
                bg=lbg.get(lh,lbg['Medium']); co=lco.get(lh,lco['Medium'])
                t_icon = type_icons.get(r.get('type',''), '⚠️')
                st.markdown(
                    "<div style='background:{bg};border:1.5px solid {co}33;border-radius:12px;"
                    "padding:0.9rem 1.05rem;margin-bottom:0.55rem;box-shadow:var(--shadow-xs);'>"
                    "<div style='display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:0.35rem;'>"
                    "<div style='display:flex;align-items:center;gap:0.5rem;'>"
                    "<span style='font-size:1rem;'>{t_icon}</span>"
                    "<span style='font-weight:700;font-size:12px;color:var(--text-primary);'>{risk}</span>"
                    "</div>"
                    "<div style='display:flex;gap:0.35rem;'>"
                    "<span style='background:{lbg_lh};color:{lco_lh};border:1.5px solid {lco_lh}55;"
                    "font-size:9px;font-weight:700;padding:0.15rem 0.5rem;border-radius:4px;"
                    "text-transform:uppercase;font-family:Syne,sans-serif;'>L:{lh}</span>"
                    "<span style='background:{lbg_im};color:{lco_im};border:1.5px solid {lco_im}55;"
                    "font-size:9px;font-weight:700;padding:0.15rem 0.5rem;border-radius:4px;"
                    "text-transform:uppercase;font-family:Syne,sans-serif;'>I:{im}</span>"
                    "<span style='background:var(--green-xlight);color:var(--green-mid);"
                    "font-size:9px;font-weight:700;padding:0.15rem 0.5rem;border-radius:4px;"
                    "font-family:Syne,sans-serif;'>{rtype}</span>"
                    "</div></div>"
                    "<div style='font-size:11px;color:var(--text-secondary);margin-bottom:0.35rem;"
                    "padding-left:1.5rem;'>{desc}</div>"
                    "<div style='font-size:11px;color:var(--text-secondary);margin-bottom:0.2rem;"
                    "padding-left:1.5rem;'>🛡️ <strong>Mitigation:</strong> {mit}</div>"
                    "<div style='font-size:10px;color:var(--text-muted);padding-left:1.5rem;'>"
                    "👤 Owner: <strong>{owner}</strong></div></div>".format(
                        bg=bg, co=co, t_icon=t_icon, risk=r.get('risk',''),
                        lbg_lh=lbg.get(lh), lco_lh=lco.get(lh), lh=lh,
                        lbg_im=lbg.get(im), lco_im=lco.get(im), im=im,
                        rtype=r.get('type',''), desc=r.get('description',''),
                        mit=r.get('mitigation',''), owner=r.get('owner','')
                    ), unsafe_allow_html=True
                )

        if st.button("📚 Show Risk Lessons from Past Projects", key="risk_lessons_btn"):
            with st.spinner("Extracting lessons from past projects..."):
                risks = st.session_state.get('current_risks', [])
                if not risks:
                    st.warning("Please extract risks first.")
                else:
                    risk_list_text = "\n".join([f"- {r.get('risk','')} ({r.get('type','')}): {r.get('description','')}" for r in risks])
                    if len(risk_list_text) > 2000: risk_list_text = risk_list_text[:2000] + "..."
                    past_patterns_text = st.session_state.past_patterns or "No patterns available"
                    if len(past_patterns_text) > 3000: past_patterns_text = past_patterns_text[:3000] + "..."
                    lesson_prompt = (
                        "Based on the following risks identified in this project:\n{risk_list}\n\n"
                        "Past project lessons and patterns:\n{past}\n\n"
                        "Write a concise set of lessons learned (bullet points) that explain how similar risks "
                        "were mitigated in past projects, and suggest specific guardrails or actions to prevent "
                        "them in this project. Keep it actionable and brief."
                    ).format(risk_list=risk_list_text, past=past_patterns_text)
                    lessons = cortex_call(lesson_prompt, "Risk Lessons")
                    if lessons:
                        st.session_state.risk_lessons = lessons; st.rerun()
                    else:
                        st.warning("Could not retrieve lessons.")

        if st.session_state.risk_lessons:
            st.markdown(
                "<div style='margin-top:0.8rem;background:var(--blue-bg);border:1.5px solid rgba(26,94,168,0.2);"
                "border-left:4px solid var(--blue);border-radius:0 12px 12px 12px;padding:1rem 1.1rem;'>"
                "<div style='font-size:9px;font-weight:800;color:var(--blue);text-transform:uppercase;"
                "letter-spacing:0.14em;margin-bottom:0.55rem;font-family:Syne,sans-serif;'>📖 Risk Lessons Learned</div>"
                "<div style='font-size:12px;color:var(--text-primary);line-height:1.8;'>"
                + st.session_state.risk_lessons.replace('\n','<br>') +
                "</div></div>",
                unsafe_allow_html=True
            )

    with tab_stakeholder:
        st.markdown("<h2>👥 Stakeholder Map & Communication Plan</h2>", unsafe_allow_html=True)

        # Full-width button
        if st.button("🤖 Generate Stakeholder Map with Cortex", key="gen_stakeholder", use_container_width=True):
            with st.spinner("Generating stakeholder map…"):
                sh_prompt = (
                    "You are a PMO consultant. For the project below, create a stakeholder engagement matrix.\n\n"
                    "Project: {pname}\nCustomer: {cust}\nType: {ptype}\nDescription: {desc}\n"
                    "Team: {team}\n\n"
                    "Return ONLY valid JSON array of stakeholders:\n"
                    '[{{"stakeholder":"Project Sponsor","role":"Executive","interest":"High",'
                    '"influence":"High","communication_freq":"Weekly","channel":"Email/Steering",'
                    '"key_concern":"ROI and timeline","engagement_strategy":"Executive steering committee"}}]\n\n'
                    "Include 8-10 stakeholders covering: sponsor, business owner, data architect, "
                    "engineers, QA, InfoSec, end users, PMO, and any project-specific roles."
                ).format(
                    pname=plan_name, cust=customer_name,
                    ptype=project_type, desc=new_project_desc[:200], team=team_roster
                )
                raw_sh = cortex_call(sh_prompt, "Stakeholder Map")
                
                # Robust JSON parsing
                stakeholders = []
                parse_error = None
                if raw_sh:
                    try:
                        clean = raw_sh.strip()
                        for fence in ["```json", "```JSON", "```"]:
                            if clean.startswith(fence):
                                clean = clean[len(fence):]
                            if clean.endswith(fence):
                                clean = clean[:-len(fence)]
                        clean = clean.strip()
                        start = clean.find('[')
                        end = clean.rfind(']')
                        if start != -1 and end != -1:
                            clean = clean[start:end+1]
                        stakeholders = json.loads(clean)
                        if not isinstance(stakeholders, list):
                            raise ValueError("Not a JSON array")
                    except Exception as e:
                        parse_error = f"Parsing failed: {e}\nRaw response:\n{raw_sh[:500]}"
                        st.error(parse_error)
                        stakeholders = []
                
                if stakeholders:
                    st.session_state.stakeholder_map = stakeholders
                else:
                    # Fallback to default stakeholder list
                    st.session_state.stakeholder_map = [
                        {"stakeholder":"Project Sponsor","role":"Executive","interest":"High","influence":"High","communication_freq":"Weekly","channel":"Email/Steering","key_concern":"Budget & ROI delivery","engagement_strategy":"Executive steering committee; weekly status"},
                        {"stakeholder":"Business Owner","role":"Senior Manager","interest":"High","influence":"High","communication_freq":"Weekly","channel":"Dashboard","key_concern":"Scope & timeline","engagement_strategy":"Weekly demo; sign-off gates"},
                        {"stakeholder":"Data Architect","role":"Technical","interest":"High","influence":"Medium","communication_freq":"Daily","channel":"Standup","key_concern":"Architecture decisions","engagement_strategy":"Daily standup; architecture reviews"},
                        {"stakeholder":"Data Engineers","role":"Technical","interest":"Medium","influence":"Medium","communication_freq":"Daily","channel":"Standup","key_concern":"Task clarity","engagement_strategy":"Sprint planning; daily standups"},
                        {"stakeholder":"QA Lead","role":"Technical","interest":"High","influence":"Medium","communication_freq":"Weekly","channel":"Jira/Email","key_concern":"Test coverage","engagement_strategy":"UAT sign-off gates; defect triage"},
                        {"stakeholder":"InfoSec","role":"Governance","interest":"Medium","influence":"High","communication_freq":"Fortnightly","channel":"Email","key_concern":"Security & compliance","engagement_strategy":"Fortnightly check-ins; RBAC approvals"},
                        {"stakeholder":"End Users","role":"Business","interest":"Medium","influence":"Low","communication_freq":"Monthly","channel":"Training","key_concern":"Ease of use","engagement_strategy":"UAT sessions; training plan"},
                        {"stakeholder":"PMO","role":"Governance","interest":"Low","influence":"Medium","communication_freq":"Weekly","channel":"Report","key_concern":"Governance & reporting","engagement_strategy":"Weekly status report; risk escalation"},
                    ]
                    st.info("⚠️ Could not parse Cortex response. Loaded default stakeholder map. Check raw response above for debugging.")
                st.rerun()

        # Display stakeholder map if available
        if st.session_state.stakeholder_map:
            stakeholders = st.session_state.stakeholder_map

            # Influence / Interest matrix
            st.markdown(
                "<div style='font-size:10px;font-weight:800;color:var(--text-muted);text-transform:uppercase;"
                "letter-spacing:0.14em;margin:0.5rem 0 0.8rem 0;font-family:Syne,sans-serif;'>"
                "📊 Influence / Interest Matrix</div>",
                unsafe_allow_html=True
            )

            quadrant_map = {"High/High":[],"High/Low":[],"Low/High":[],"Low/Low":[]}
            for s in stakeholders:
                inf = s.get('influence','Medium').strip().title()
                itr = s.get('interest','Medium').strip().title()
                if inf == "High" and itr == "High": quadrant_map["High/High"].append(s.get('stakeholder',''))
                elif inf == "High" and itr == "Low": quadrant_map["High/Low"].append(s.get('stakeholder',''))
                elif inf == "Low" and itr == "High": quadrant_map["Low/High"].append(s.get('stakeholder',''))
                else: quadrant_map["Low/Low"].append(s.get('stakeholder',''))

            q_colors = {
                "High/High": ("#fdecea","#c0392b","Manage Closely"),
                "High/Low":  ("#e3eeff","#1a5ea8","Keep Satisfied"),
                "Low/High":  ("#e4f5ec","#196b41","Keep Informed"),
                "Low/Low":   ("#f5f5f5","#888","Monitor"),
            }
            qc1, qc2 = st.columns(2)
            quadrants = list(quadrant_map.items())
            for i, (qk, names) in enumerate(quadrants):
                bg, col, label = q_colors[qk]
                with (qc1 if i % 2 == 0 else qc2):
                    bullets = "".join(f"<div style='font-size:11px;color:{col};padding:0.15rem 0;font-weight:600;'>• {n}</div>" for n in names) if names else "<div style='font-size:11px;color:#bbb;font-style:italic;'>None</div>"
                    st.markdown(
                        f"<div style='background:{bg};border:1.5px solid {col}33;border-radius:12px;"
                        f"padding:0.9rem 1rem;margin-bottom:0.65rem;'>"
                        f"<div style='font-size:9px;font-weight:800;color:{col};text-transform:uppercase;"
                        f"letter-spacing:0.12em;margin-bottom:0.5rem;font-family:Syne,sans-serif;'>"
                        f"{'High' if 'High/H' in qk else 'Low'} Influence / {'High' if qk.endswith('High') else 'Low'} Interest"
                        f" — {label}</div>"
                        f"{bullets}</div>",
                        unsafe_allow_html=True
                    )

            st.markdown("<hr>", unsafe_allow_html=True)

            # Detailed table
            st.markdown(
                "<div style='font-size:10px;font-weight:800;color:var(--text-muted);text-transform:uppercase;"
                "letter-spacing:0.14em;margin-bottom:0.75rem;font-family:Syne,sans-serif;'>"
                "📋 Full Stakeholder Register</div>",
                unsafe_allow_html=True
            )
            int_colors = {"High":"#196b41","Medium":"#c8922a","Low":"#888888"}
            int_bgs    = {"High":"#e4f5ec","Medium":"#fef7e8","Low":"#f5f5f5"}

            for s in stakeholders:
                inf = s.get('influence','Medium').strip().title()
                itr = s.get('interest','Medium').strip().title()
                inf_col = int_colors.get(inf,"#888"); inf_bg = int_bgs.get(inf,"#f5f5f5")
                itr_col = int_colors.get(itr,"#888"); itr_bg = int_bgs.get(itr,"#f5f5f5")
                st.markdown(
                    "<div style='background:#ffffff;border:1.5px solid var(--border);border-radius:12px;"
                    "padding:0.85rem 1.1rem;margin-bottom:0.5rem;box-shadow:var(--shadow-xs);'>"
                    "<div style='display:flex;align-items:center;gap:0.75rem;margin-bottom:0.5rem;flex-wrap:wrap;'>"
                    "<span style='font-size:13px;font-weight:800;color:var(--text-primary);"
                    "font-family:Syne,sans-serif;'>{name}</span>"
                    "<span style='font-size:10px;color:var(--text-muted);font-weight:500;'>{role}</span>"
                    "<span style='background:{inf_bg};color:{inf_col};border:1.5px solid {inf_col}44;"
                    "font-size:9px;font-weight:700;padding:0.1rem 0.5rem;border-radius:4px;"
                    "font-family:Syne,sans-serif;'>Influence: {inf}</span>"
                    "<span style='background:{itr_bg};color:{itr_col};border:1.5px solid {itr_col}44;"
                    "font-size:9px;font-weight:700;padding:0.1rem 0.5rem;border-radius:4px;"
                    "font-family:Syne,sans-serif;'>Interest: {itr}</span>"
                    "<span style='font-size:9px;color:var(--text-muted);margin-left:auto;'>"
                    "📅 {freq} · 📡 {ch}</span>"
                    "</div>"
                    "<div style='display:flex;gap:1.5rem;flex-wrap:wrap;'>"
                    "<div style='font-size:11px;color:var(--text-secondary);'>"
                    "<strong style='color:var(--text-primary);'>Key Concern:</strong> {concern}</div>"
                    "<div style='font-size:11px;color:var(--text-secondary);'>"
                    "<strong style='color:var(--text-primary);'>Strategy:</strong> {strategy}</div>"
                    "</div></div>".format(
                        name=s.get('stakeholder',''), role=s.get('role',''),
                        inf=inf, inf_col=inf_col, inf_bg=inf_bg,
                        itr=itr, itr_col=itr_col, itr_bg=itr_bg,
                        freq=s.get('communication_freq','—'), ch=s.get('channel','—'),
                        concern=s.get('key_concern','—'), strategy=s.get('engagement_strategy','—')
                    ), unsafe_allow_html=True
                )
        else:
            st.info("Click the button above to generate an AI-powered stakeholder engagement plan for this project.")

    with tab_roi:
        st.markdown("<h2>💰 ROI & Business Case Report</h2>", unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:12px;color:var(--text-muted);margin-bottom:1.2rem;'>"
            "AI-generated ROI analysis for <strong style='color:var(--text-primary);'>{}</strong> · {}</div>".format(
                plan_name, datetime.now().strftime('%d %B %Y')
            ), unsafe_allow_html=True
        )

        kpi_c1, kpi_c2, kpi_c3, kpi_c4, kpi_c5 = st.columns(5)
        with kpi_c1: st.metric("ROI", "{}%".format(roi_data['roi_percentage']), "Return on investment")
        with kpi_c2: st.metric("Total Value", "£{:,}".format(roi_data['total_roi_value']), "Per project")
        with kpi_c3: st.metric("Net Benefit", "£{:,}".format(roi_data['net_benefit']), "After £{:,} AI cost".format(roi_data['ai_planning_cost']))
        with kpi_c4: st.metric("Time Saved", "{}%".format(roi_data['time_saved_pct']), "{}h saved".format(roi_data['manual_hours'] - roi_data['cortex_hours']))
        with kpi_c5: st.metric("Risk Avoided", "£{:,}".format(roi_data['risk_saving']), "{} risks mitigated".format(roi_data['risks_identified']))

        st.markdown("<hr>", unsafe_allow_html=True)
        roi_tab_l, roi_tab_r = st.columns([3,2])

        with roi_tab_l:
            st.markdown(
                "<div style='font-size:10px;font-weight:800;color:var(--green-mid);text-transform:uppercase;"
                "letter-spacing:0.12em;margin-bottom:0.7rem;font-family:Syne,sans-serif;'>📊 Cost-Benefit Breakdown</div>",
                unsafe_allow_html=True
            )

            breakdown_rows = [
                ("COSTS","","",""),
                ("AI-Assisted Planning","{} hours".format(roi_data['cortex_hours']),"@ £{}/hr".format(roi_data.get('hourly_rate',80)),"£{:,}".format(roi_data['ai_planning_cost'])),
                ("Total Investment","","","£{:,}".format(roi_data['ai_planning_cost'])),
                ("BENEFITS","","",""),
                ("Planning Time Saved","{} hours".format(roi_data['manual_hours'] - roi_data['cortex_hours']),"@ £{}/hr".format(roi_data.get('hourly_rate',80)),"£{:,}".format(roi_data['planning_saving'])),
                ("Risk Incident Avoidance","{} risks × {} weeks".format(roi_data['risks_identified'], roi_data['duration_weeks']),"£500/week per risk","£{:,}".format(roi_data['risk_saving'])),
                ("Total Benefits","","","£{:,}".format(roi_data['total_roi_value'])),
                ("NET BENEFIT","","","£{:,}".format(roi_data['net_benefit'])),
                ("ROI %","","","{}%".format(roi_data['roi_percentage'])),
            ]

            table_html = (
                "<div style='overflow-x:auto;border-radius:12px;border:1.5px solid var(--border);"
                "box-shadow:var(--shadow-sm);'><table style='width:100%;border-collapse:collapse;'><thead>"
            )
            for h in ["Item","Detail","Rate / Basis","Value"]:
                table_html += (
                    "<th style='padding:0.6rem 0.9rem;text-align:left;font-size:9px;font-weight:800;"
                    "letter-spacing:0.09em;color:var(--green-mid);background:var(--green-xlight);"
                    "border-bottom:2px solid var(--border-strong);font-family:Syne,sans-serif;'>{}</th>".format(h)
                )
            table_html += "</thead><tbody>"

            for idx, (item, detail, basis, value) in enumerate(breakdown_rows):
                is_section = item in {"COSTS","BENEFITS"}
                is_total   = item in {"Total Investment","Total Benefits","NET BENEFIT","ROI %"}
                row_bg = "var(--green-xlight)" if is_section else ("rgba(25,107,65,0.06)" if is_total else ("#ffffff" if idx%2==0 else "var(--bg-panel-alt)"))
                font_w = "800" if is_section or is_total else "500"
                txt_col = "var(--green-mid)" if is_section else "var(--text-primary)"
                val_col = "var(--green-deep)" if (value.startswith("£") and not is_section) else "var(--green-mid)" if is_section else "var(--text-muted)"
                font_family = "font-family:Syne,sans-serif;" if is_section or is_total else ""

                table_html += "<tr style='background:{};'>".format(row_bg)
                table_html += "<td style='padding:0.5rem 0.9rem;font-size:11px;font-weight:{};color:{};border-bottom:1px solid var(--border);{}'>{}</td>".format(font_w, txt_col, font_family, item)
                table_html += "<td style='padding:0.5rem 0.9rem;font-size:11px;color:var(--text-secondary);border-bottom:1px solid var(--border);'>{}</td>".format(detail)
                table_html += "<td style='padding:0.5rem 0.9rem;font-size:11px;color:var(--text-muted);border-bottom:1px solid var(--border);'>{}</td>".format(basis)
                table_html += "<td style='padding:0.5rem 0.9rem;font-size:11px;font-weight:700;color:{};border-bottom:1px solid var(--border);text-align:right;'>{}</td>".format(val_col, value)
                table_html += "</tr>"

            table_html += "</tbody></table></div>"
            st.markdown(table_html, unsafe_allow_html=True)

            st.markdown("""
            <div style='margin-top:1rem;background:var(--gold-bg);border:1.5px solid rgba(200,146,42,0.22);
                        border-left:4px solid var(--gold);border-radius:0 12px 12px 0;padding:0.85rem 1rem;'>
              <div style='font-size:9px;font-weight:800;color:var(--gold);text-transform:uppercase;
                          letter-spacing:0.14em;margin-bottom:0.4rem;font-family:Syne,sans-serif;'>⚠️ Assumptions & Notes</div>
              <div style='font-size:11px;color:var(--text-secondary);line-height:1.8;'>
                • Hourly rate basis: £{hrly}/hr (PM / SA grade blended rate)<br>
                • Manual planning baseline: {manual_hours} hours (PM+SA)<br>
                • AI-assisted planning: {cortex_hours} hour(s) (SOW upload + generation + review)<br>
                • Risk savings: £500 per week per risk identified (prevents £1,000/week avg issue cost)<br>
                • ROI calculated on direct planning cost savings + risk avoidance value only<br>
                • Excludes: accelerated time-to-market, quality uplift, team capability development
              </div>
            </div>
            """.format(
                hrly=roi_data.get('hourly_rate',80),
                manual_hours=roi_data['manual_hours'],
                cortex_hours=roi_data['cortex_hours']
            ), unsafe_allow_html=True)

        with roi_tab_r:
            st.markdown(
                "<div style='font-size:10px;font-weight:800;color:var(--green-mid);text-transform:uppercase;"
                "letter-spacing:0.12em;margin-bottom:0.7rem;font-family:Syne,sans-serif;'>📈 Value Visualisation</div>",
                unsafe_allow_html=True
            )

            roi_colour = "#15803d" if roi_data['roi_percentage'] >= 200 else "#196b41" if roi_data['roi_percentage'] >= 100 else "#c8922a"
            st.markdown("""
            <div style="background:linear-gradient(135deg,#f0fdf4,#dcfce7);border:1.5px solid rgba(34,197,94,0.25);
                        border-radius:16px;padding:1.3rem 1.5rem;margin-bottom:0.9rem;
                        box-shadow:var(--shadow-green);text-align:center;position:relative;overflow:hidden;">
              <div style="position:absolute;top:0;left:0;right:0;height:2.5px;
                          background:linear-gradient(90deg,#0d4428,#196b41,#30c472,#c8922a);"></div>
              <div style="font-size:9px;font-weight:800;color:#166534;text-transform:uppercase;
                          letter-spacing:0.18em;margin-bottom:0.4rem;font-family:Syne,sans-serif;">Return on Investment</div>
              <div style="font-size:52px;font-weight:900;color:{roi_colour};line-height:1;
                          letter-spacing:-0.03em;font-family:Syne,sans-serif;">{roi_percent}%</div>
              <div style="font-size:11px;color:#3d6b4a;margin-top:0.45rem;font-weight:500;">per project cycle</div>
            </div>
            """.format(roi_colour=roi_colour, roi_percent=roi_data['roi_percentage']), unsafe_allow_html=True)

            waterfall_items = [
                ("AI Investment", -roi_data['ai_planning_cost'], "#c0392b", "#fdecea"),
                ("Planning Savings", roi_data['planning_saving'], "#196b41", "#e4f5ec"),
                ("Risk Avoidance", roi_data['risk_saving'], "#1a5ea8", "#e3eeff"),
                ("Net Benefit", roi_data['net_benefit'], "#0d4428" if roi_data['net_benefit'] >= 0 else "#c0392b", "#f0fdf4"),
            ]
            max_abs = max(abs(x[1]) for x in waterfall_items) or 1
            for label, val, color, bg in waterfall_items:
                bar_w = int(abs(val) / max_abs * 100)
                val_str = "£{:,}".format(abs(val))
                is_neg  = val < 0
                st.markdown("""
<div style="margin-bottom:0.6rem;">
  <div style="display:flex;justify-content:space-between;font-size:10px;font-weight:600;
              color:var(--text-secondary);margin-bottom:0.22rem;">
    <span>{label}</span>
    <span style="color:{color};font-weight:700;">{sign}{val_str}</span>
  </div>
  <div style="height:24px;background:var(--border);border-radius:8px;overflow:hidden;position:relative;">
    <div style="position:absolute;{align}:0;height:100%;width:{bar_w}%;background:{bg};
                border:1px solid {color}33;border-radius:8px;display:flex;align-items:center;{justify}">
      <span style="font-size:10px;font-weight:700;color:{color};white-space:nowrap;
                   font-family:Syne,sans-serif;">{sign}{val_str}</span>
    </div>
  </div>
</div>
""".format(
                    label=label, color=color,
                    sign='−' if is_neg else '+', val_str=val_str,
                    align='right' if is_neg else 'left',
                    bar_w=bar_w, bg=bg,
                    justify='justify-content:flex-end;padding-right:8px;' if is_neg else 'justify-content:flex-start;padding-left:8px;'
                ), unsafe_allow_html=True)

            st.markdown("""
            <div style="background:#ffffff;border:1.5px solid var(--border);border-radius:12px;
                        padding:0.9rem 1rem;margin-top:0.7rem;box-shadow:var(--shadow-xs);">
              <div style="font-size:9px;font-weight:800;color:var(--green-mid);text-transform:uppercase;
                          letter-spacing:0.16em;margin-bottom:0.55rem;font-family:Syne,sans-serif;">⏱ Time Comparison</div>
              <div style="display:flex;justify-content:space-between;align-items:center;
                          font-size:11px;padding:0.28rem 0;border-bottom:1px solid var(--border);">
                <span style="color:var(--text-muted);">Traditional approach</span>
                <span style="color:var(--red);font-weight:700;">{manual_hours} hours</span>
              </div>
              <div style="display:flex;justify-content:space-between;align-items:center;
                          font-size:11px;padding:0.28rem 0;border-bottom:1px solid var(--border);">
                <span style="color:var(--text-muted);">AI-assisted approach</span>
                <span style="color:var(--green-mid);font-weight:700;">{cortex_hours} hour(s)</span>
              </div>
              <div style="display:flex;justify-content:space-between;align-items:center;
                          font-size:11px;padding:0.28rem 0;">
                <span style="color:var(--text-muted);">Time reduction</span>
                <span style="color:var(--green-deep);font-weight:800;font-family:Syne,sans-serif;">{time_saved}% faster</span>
              </div>
            </div>
            """.format(
                manual_hours=roi_data['manual_hours'],
                cortex_hours=roi_data['cortex_hours'],
                time_saved=roi_data['time_saved_pct']
            ), unsafe_allow_html=True)

    with tab_export:
        st.markdown("<h2>📤 Export & Save</h2>", unsafe_allow_html=True)
        ec1,ec2,ec3 = st.columns(3)
        with ec1:
            st.markdown("<div class='plan-section-title'>💾 Save to Snowflake</div>", unsafe_allow_html=True)
            save_status = st.selectbox("Status", ["In Progress","Planning","On Hold"])
            st.markdown(
                "<div style='background:var(--gold-bg);border:1.5px solid rgba(200,146,42,0.28);"
                "border-left:3px solid var(--gold);border-radius:10px;padding:0.75rem 0.9rem;margin-bottom:0.65rem;'>"
                "<div style='font-size:9px;font-weight:700;color:var(--gold);text-transform:uppercase;"
                "margin-bottom:0.35rem;font-family:Syne,sans-serif;'>⚠️ Review Gate</div>"
                "<div style='font-size:11px;color:var(--text-secondary);'>"
                "Has the Project Plan been reviewed before saving?</div></div>",
                unsafe_allow_html=True
            )
            review_confirmed = st.radio(
                "Plan Reviewed?",
                options=["— Select —","✅ Yes — Plan reviewed, save it","❌ No — Not yet reviewed"],
                index=0, key="save_review_gate", label_visibility="collapsed"
            )
            if review_confirmed == "✅ Yes — Plan reviewed, save it":
                if st.button("💾 Save to Projects Table", key="save_btn"):
                    try:
                        esc=(plan or '').replace("'","''"); pn=plan_name.replace("'","''")
                        wbs_part=esc[:3000]; depl_part=esc[3000:5000] if len(esc)>3000 else ''
                        session.sql(f"INSERT INTO projects (project_name,description,status,wbs_summary,risk_log_summary,test_cases_summary,deployment_plan) VALUES('{pn}','AI-generated plan — reviewed','{save_status}','{wbs_part}','','','{depl_part}')").collect()
                        st.success("✅ Project saved to DB.")
                        st.session_state.past_projects_df = load_past_projects()
                        st.session_state.imported_csv_names.add(plan_name)
                        st.rerun()
                    except Exception as e: st.error(f"Save failed: {e}")
            elif review_confirmed == "❌ No — Not yet reviewed":
                st.markdown(
                    "<div style='background:var(--red-bg);border:1.5px solid rgba(192,57,43,0.28);"
                    "border-radius:10px;padding:0.65rem 0.9rem;'>"
                    "<div style='font-size:11px;color:var(--red);font-weight:700;'>🚫 Save blocked.</div>"
                    "<div style='font-size:11px;color:var(--text-secondary);'>Please review before saving.</div>"
                    "</div>",
                    unsafe_allow_html=True
                )
        with ec2:
            st.markdown("<div class='plan-section-title'>📄 Download Text</div>", unsafe_allow_html=True)
            st.download_button(
                "⬇ Download .txt",
                data=plan or "",
                file_name="{}_plan.txt".format(plan_name.replace(' ','_')),
                mime="text/plain"
            )
        with ec3:
            st.markdown("<div class='plan-section-title'>📊 Download Excel (6 Sheets)</div>", unsafe_allow_html=True)
            st.markdown(
                "<div style='font-size:11px;color:var(--text-muted);margin-bottom:0.5rem;'>"
                "Summary · WBS/Sprints · RAID Log · Test Cases · Timeline · Stakeholder Matrix</div>",
                unsafe_allow_html=True
            )
            if st.button("📎 Build Excel Workbook"):
                with st.spinner("Building Excel…"):
                    sp = sprints if sprints else parse_wbs_into_sprints(plan, 12, meth)
                    excel_bytes = build_excel(plan, plan_name, sp, meth)
                st.download_button(
                    label="⬇ Download .xlsx",
                    data=excel_bytes,
                    file_name="{}_ProjectPlan.xlsx".format(plan_name.replace(' ','_')),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        with st.expander("🔍 Preview plan text"):
            safe_plan = (plan or "").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            st.markdown("<div class='plan-preview-wrap'><pre>" + safe_plan + "</pre></div>", unsafe_allow_html=True)

elif not generate_button:
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align:center;padding:3.5rem 1rem;">
        <div style="font-size:2.5rem;margin-bottom:0.75rem;">📊</div>
        <div style="font-size:16px;font-weight:800;color:var(--text-secondary);margin-bottom:0.5rem;
                    font-family:Syne,sans-serif;">Ready to generate your project plan</div>
        <div style="font-size:12px;color:var(--text-muted);max-width:440px;margin:0 auto;line-height:1.85;">
            Upload your SOW · Fill in project details · Click
            <strong style="color:var(--green-mid);">Generate Project Plan</strong><br/><br/>
            Cortex will scan institutional memory and produce:<br/>
            WBS · Sprint Board · RAID Log · Test Cases · Timeline · Deployment Strategy · Stakeholder Map
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown("""
<div class='footer'>
  Powered by <span>Snowflake Cortex</span> &nbsp;·&nbsp;
  PMO Intelligence &nbsp;·&nbsp;
  Institutional Memory → Faster Delivery &nbsp;·&nbsp;
  <span>COCO 2026</span>
</div>
""", unsafe_allow_html=True)
